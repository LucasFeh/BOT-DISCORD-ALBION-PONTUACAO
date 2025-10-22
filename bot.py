
import datetime
from email.policy import default
import re
import aiofiles
import pytz
import re
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import csv
import discord
from discord.ext import commands
from discord import app_commands
import aiohttp
import json
import os
import asyncio


intents = discord.Intents.default()
def ler_token():
    try:
        with open("TOKEN.TXT", "r", encoding="utf-8") as f:
            return f.read().strip()
    except Exception:
        return None
token = ler_token()
intents.message_content = True
intents.members = True  # <-- Adicione esta linha!
intents = discord.Intents.all()
print("🧠 Intents.members está:", intents.members)
bot = commands.Bot(command_prefix='!', intents=intents)

# ID fixo da guilda LOUCOS POR PVE (apenas para API do Albion Online)
GUILD_ID = "QDufxXRfSiydcD58_Lo9KA"

# Dicionário de pontuação por conteúdo
PONTOS_POR_CONTEUDO = {
    "DG BENEFICIENTE": 20,
    "ARMA 4.4": 40,
    "PRATA (4M)": 70,
    "ARMA 8.3": 300,
    "ARMA 8.4": 5500,
    # você pode adicionar mais tipos depois
}

TIPOS_DE_DG = {
    "SORTEIO",
    "PATROCIONADOR",
    "PONTUAÇÃO",
    "RECRUTADOR"
}

# Mapear ícones
icones = {
    "MONTARIA - (600k)": "🐎",
    "RE-GEAR (4M)": "🛡️",
    "ARMA 4.4": "🗡️",
    "ARMA 8.4": "🗡️",
    "MONTARIA (1.5M)": "🐎",
    "MONTARIA (4M)": "🐎",
    "PRATA (4M)": "💰",
    "ARMA 8.3": "🗡️",
    "MAMUTE": "🐘",
    "SORTEIO": "🎲",
    "PATROCIONADOR": "🎁",
    "PONTUAÇÃO": "🏆",
    "RECRUTADOR": "🤝"
}
    

# Arquivo JSON para armazenar a pontuação (no mesmo diretório do bot.py)

ARQUIVO_PONTUACAO = "pontuacaoMembros.json"
ARQUIVO_SORTEIOS = "sorteios.json"
ARQUIVO_PATROCINADOR = "patrocinadores.json"


# ;
# ;
# ;
# ;
# ;
# -------------------------------------------------- SETUP() -----------------------------------------------------
# ;
# ;
# ;
# ;
# ;

# Adiciona tarefa agendada para atualizar patrocinadores todo domingo à meia-noite
@bot.event
async def on_ready():
    await atualizar_ranking()
    print(f'✅ Bot conectado como {bot.user}')
    try:
        synced = await bot.tree.sync()
        print(f"🔗 {len(synced)} comandos sincronizados (Slash Commands).")
    except Exception as e:
        print(f"❌ Erro ao sincronizar comandos: {e}")

    # Inicia a tarefa agendada
    if not hasattr(bot, 'patrocinador_task_started'):
        bot.loop.create_task(agendar_atualizacao_patrocinadores())
        bot.patrocinador_task_started = True

# ;
# ;
# ;
# ;
# ;
# ----------------------------------------- ADICONAR GANHADORES SORTEIO ---------------------------------
# ;
# ;
# ;
# ;
# ;

# Observa mensagens no canal ・dg-beneficente e adiciona ganhadores ao sorteios.json
@bot.event
async def on_message(message):
    # Ignorar mensagens do próprio bot (mas permitir mensagens de outros bots)
    if message.author.id == bot.user.id:
        return

    print(f"[DEBUG] Mensagem recebida no canal: {message.channel.name}")

    # Verifica se é o canal correto (pode ser pelo nome ou ID)
    if message.channel.name == "🎁・dg-beneficente":
        texto = message.content
        # print(f"[DEBUG] Conteúdo da mensagem: {texto}")
        # Regex para pegar nomes no formato @[TAG] Nome
        padrao = r"@\[([^\]]+)\]\s*([^,\n!]+)"
        nomes = re.findall(padrao, texto)
        # print(f"[DEBUG] Nomes extraídos: {nomes}")
        if nomes and "You won the DG BENEFICENTE!" in texto:
            ganhadores = []
            for tag, nome in nomes:
                nome_completo = f"[{tag}] {nome.strip()}"
                # print(f"[DEBUG] Adicionando ao sorteio: {nome_completo}")
                adicionar_sorteio(nome_completo)
                ganhadores.append(nome_completo)
            if ganhadores:
                nomes_str = ", ".join(f"**{n}**" for n in ganhadores)
                msg = (
                    f"🎉 Parabéns {nomes_str}!\n"
                    "Adicionado na lista de DG beneficente, já pode ir lá jogar meus queridos!! 🥳💥"
                )
                await message.channel.send(msg)
    # Permite que comandos normais funcionem
    await bot.process_commands(message)


# ;
# ;
# ;
# ;
# ;
# ------------------------------------- LISTAR RECOMPENSAS DE PONTOS ---------------------------------
# ;
# ;
# ;
# ;
# ;
    
@bot.tree.command(name="recompensas", description="Mostra a tabela de recompensas")
async def recompensas(interaction: discord.Interaction):

    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar


    embed = discord.Embed(
        title="📊 TABELA DE RECOMPENSAS - LOUCOS POR PVE [PVE]",
        description="Sistema de recompensas para membros da guilda.\n\n",
        color=0x00ff00
    )

    for tipo, pontos in PONTOS_POR_CONTEUDO.items():
        icone = icones.get(tipo, "📋")
        nome_formatado = f"{icone} {tipo.replace('-', ' ').title()}"
        
        # Cada linha é 1 field com nome e pontos
        embed.add_field(
            name=nome_formatado,
            value=f"```ansi\n\u001b[36m{pontos} pts\u001b[0m```",
            inline=False  # inline=False faz cada linha ocupar toda a largura do embed
        )

    embed.set_footer(text="Use !conteudo <caller> <tipo> <participantes> para registrar")
    await interaction.response.send_message(embed=embed)

# cria as opções automaticamente a partir do dicionário
TIPOS_CHOICES = [
    app_commands.Choice(name=nome, value=nome)
    for nome in TIPOS_DE_DG
]

# ;
# ;
# ;
# ;
# ;
# ------------------------------------- FUNÇÃO PRINCIPAL (CONTEUDINHO ) ---------------------------------
# ;
# ;
# ;
# ;
# ;

conteudos_em_aberto = {}

@bot.tree.command(name="dg_beneficente", description="Registra uma DG beneficente")
@app_commands.describe(
    tipo="Tipo da Beneficiente",
    integrantes="Lista de integrantes separados por espaço"
)
@app_commands.choices(tipo=[
    app_commands.Choice(name=f"{icones.get(key, '📋')} {key.replace('-', ' ').title()}", value=key)
    for key in TIPOS_DE_DG
])
async def dg_beneficente(
    interaction: discord.Interaction,
    tipo: app_commands.Choice[str],
    integrantes: str
):
    

    if await permitir_comando_apenas_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, não está no canal permitido e o comando deve parar

    try:
        await safe_defer(interaction)
    except Exception as e:
        print(f"[WARN] Não foi possível deferir interação (/dg_beneficente): {e}")

    # Usar ID do usuário como chave única
    user_id = interaction.user.id
    usuario_comando = interaction.user.display_name  
    
    # Verificar se o usuário já tem um comando em andamento
    if user_id in conteudos_em_aberto:
        embed_erro = discord.Embed(
            title="❌ Comando em Andamento",
            description="Você já tem um comando /dg_beneficente em andamento. Finalize-o antes de iniciar outro.",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return
    
    tipo_valor = tipo.value

    # ... resto da validação permanece igual ...
    # Validar que todos os integrantes fornecidos são mentions no formato <@123> ou <@!123>
    invalid_tokens = []
    mention_pattern = re.compile(r"^<@!?(\d+)>$")
    if integrantes:
        for parte in integrantes.split():
            if not mention_pattern.match(parte):
                invalid_tokens.append(parte)

    if invalid_tokens:
        embed_erro = discord.Embed(
            title="❌ Formato inválido - Integrantes",
            description=(
                "Os participantes devem ser informados como mentions do Discord.\n"
                "Por favor, marque cada participante usando `@Nickname` (o bot recebe o formato interno `<@USERID>`).\n\n"
                f"Tokens inválidos: {', '.join(invalid_tokens)}"
            ),
            color=0xff0000
        )
        embed_erro.add_field(name="💡 Como corrigir", value="Use o @ para mencionar cada jogador; ex: `@Klartz`.", inline=False)
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    # Depois de validar mentions, converter para nomes limpos e checar self-add
    for parte in integrantes.split():
        nome_limpo = await tratar_mention(interaction, parte)
        if nome_limpo == usuario_comando:
            embed_erro = discord.Embed(
                title="❌ Erro de Integrantes",
                description=f"**{usuario_comando}**, você não pode se adicionar como integrante, você já é o caller!",
                color=0xff0000
            )
            embed_erro.add_field(name="Observação", value="**O caller já é automaticamente adicionado como participante, mas não recebe pontuação!!!**", inline=False)
            embed_erro.set_footer(text=f"Consulte /tutorial_dg para mais informações.")
            await interaction.followup.send(embed=embed_erro, ephemeral=True)
            return

    # Verificar se o tipo existe no dicionário
    if tipo_valor not in TIPOS_DE_DG:
        embed_erro = discord.Embed(
            title="❌ Tipo inválido",
            description=f"O tipo **{tipo_valor}** não foi encontrado no sistema.",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    # 🔥 NOVAS VERIFICAÇÕES POR TIPO 🔥
    member = interaction.user
    
    if tipo_valor == "PATROCIONADOR":
        with open(ARQUIVO_PATROCINADOR, 'r', encoding='utf-8') as f:
            patrocinadores_pendentes = json.load(f)

        if not verificar_tag_discord(member, "patrocinador"):
            embed_erro = discord.Embed(
                title="❌ Acesso Negado - Patrocinador",
                description="Iiiih amigo, você não tem TAG de patrocinador, procure um **BRAÇO DIREITO**, ou o **Líder da guild** para saber mais sobre ser um patrocinador.",
                color=0xff0000
            )
            embed_erro.set_footer(text="💡 Apenas membros com TAG de 'Patrocinador' podem usar este tipo.")
            await interaction.followup.send(embed=embed_erro)
            return
        if usuario_comando not in patrocinadores_pendentes:
            embed_erro = discord.Embed(
                title="❌ Acesso Negado - Patrocinador",
                description="Vissh você já fez a DG beneficiente essa semana, tente usar a sua pontuação ou espere até a próxima semana.",
                color=0xff0000
            )
            embed_erro.set_footer(text="💡 Apenas patrocinadores que ainda não fizeram a DG dessa semana podem usar este tipo.")
            await interaction.followup.send(embed=embed_erro)
            return
    elif tipo_valor == "SORTEIO":
        if not verificar_sorteio(usuario_comando):
            embed_erro = discord.Embed(
                title="❌ Acesso Negado - Sorteio",
                description="Vissh você não ganhou nenhum sorteio atualmente, sinto muito, tente usar a sua pontuação.",
                color=0xff0000
            )
            embed_erro.set_footer(text="💡 Apenas quem ganhou sorteios recentes pode usar este tipo.")
            await interaction.followup.send(embed=embed_erro)
            return
    
    elif tipo_valor == "RECRUTADOR":
        if not verificar_tag_discord(member, "recrutador"):
            embed_erro = discord.Embed(
                title="❌ Acesso Negado - Recrutador",
                description="Tá tentando usar privilégio que não é pro seu bico né?? Tente ganhar um sorteio ou use seus pontos.",
                color=0xff0000
            )
            embed_erro.set_footer(text="💡 Apenas membros com TAG de 'Recrutador' podem usar este tipo.")
            await interaction.followup.send(embed=embed_erro)
            return
    
    elif tipo_valor == "PONTUAÇÃO":
        pontos_necessarios = 10
        pontos_atuais = obter_pontuacao(usuario_comando)
        
        if pontos_atuais < pontos_necessarios:
            embed_erro = discord.Embed(
                title="❌ Pontos Insuficientes",
                description=f"**{usuario_comando}**, você não tem pontos suficientes.\n\n"
                           f"**Necessário para DG beneficiente:** {pontos_necessarios} pontos\n"
                           f"**Você tem:** {pontos_atuais} pontos",
                color=0xff0000
            )
            embed_erro.add_field(
                name="💡 Como conseguir pontos:",
                value="• Participe de DGs como Tank/Healer (+2 pts)\n• Participe de DGs como DPS (+1 pt)\n• Ganhe sorteios da guild",
                inline=False
            )
            embed_erro.set_footer(text="Use /ranking para ver o ranking de pontuação.")
            await interaction.followup.send(embed=embed_erro)
            return
         
    membros = [usuario_comando]
    if integrantes:
        for parte in integrantes.split():
            nome_limpo = await tratar_mention(interaction, parte)
            membros.append(nome_limpo)

    # Armazenar dados específicos para este usuário
    conteudos_em_aberto[user_id] = {
        "caller": usuario_comando,
        "tipo": tipo_valor,
        "membros": membros,
    }

    # Embed inicial
    icone = icones.get(tipo_valor, "📋")
    embed = discord.Embed(
        title=f"📊 PRÉVIA DE PONTUAÇÃO",
        description=f"{icone} **{tipo.name}**\n\nClique nos botões abaixo para definir Tank e Healer.\nQuando terminar, clique em **Finalizar**.",
        color=0xffa500
    )
    
    for membro in membros:
        embed.add_field(
            name=f"⚔️ {membro}",
            value="Função: **DPS**",
            inline=False
        )
    
    # Passar o user_id para a view
    view = FuncoesEquipeView(membros, interaction.user, user_id)
    view.interaction = interaction

    message = await interaction.followup.send(embed=embed, view=view)
    view.message = message
    

# ;
# ;
# ;
# ;
# ;
# ------------------------------------- SPLIT DE VALOR ---------------------------------
# ;
# ;
# ;
# ;
# ;


@bot.tree.command(name="split", description="Divide um valor entre várias pessoas")
@app_commands.describe(
    valor="Valor total (ex: 17M, 500K, 2.5B)",
    quantidade_de_membros="Número de pessoas para dividir"
)
async def split(interaction: discord.Interaction, valor: str, quantidade_de_membros: int):

    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar


    try:
        # Converter valor abreviado para número
        valor_original = valor  # Guardar o valor original para exibir
        valor_numerico = converter_valor_abreviado(valor)
        
        if quantidade_de_membros <= 0:
            raise ValueError("A quantidade deve ser maior que zero.")

        valor_por_pessoa = valor_numerico / quantidade_de_membros

        # Formatar valores para exibição
        valor_formatado = formatar_valor_abreviado(valor_numerico)
        valor_pessoa_formatado = formatar_valor_abreviado(valor_por_pessoa)

        embed = discord.Embed(
            title=f"💰 SPLIT DE VALOR",
            description=f"💰 **{valor_formatado}** dividido por **{quantidade_de_membros}** pessoas",
            color=0xffa500  # Laranja para prévia
        )

        # Campo com resumo
        embed.add_field(
            name="Resumo",
            value=f"**Valor por pessoa**: {valor_pessoa_formatado}",
            inline=False
        )

        embed.set_footer(text="Valores em formato abreviado (K=mil, M=milhão, B=bilhão) cada pessoa deve receber o valor indicado")
        
        await interaction.response.send_message(embed=embed)

    except ValueError as e:
        embed_erro = discord.Embed(
            title="❌ Erro",
            description=f"Erro ao processar o comando: {e}\n\n💡 **Formatos aceitos:**\n`17M` (17 milhões)\n`500K` (500 mil)\n`2.5B` (2.5 bilhões)\n`1000` (número normal)",
            color=0xff0000
        )
        await interaction.response.send_message(embed=embed_erro)

# ;
# ;
# ;
# ;
# ;
# ------------------------------------- MOSTRAR INFORMAÇÕES  ---------------------------------
# ;
# ;
# ;
# ;
# ;

@bot.tree.command(name="guilda", description="Mostra informações da guilda LOUCOS POR PVE")
async def guilda(interaction: discord.Interaction):
    embed_loading = discord.Embed(
        title="🔍 Buscando informações da guilda...",
        description="Consultando API do Albion Online",
        color=0xffa500
    )
    await interaction.response.send_message(embed=embed_loading)
    
    try:
        guilda_info = await buscar_guilda_por_id(GUILD_ID)
        if not guilda_info:
            embed_erro = discord.Embed(
                title="❌ Erro ao carregar dados",
                description="Não foi possível carregar as informações da guilda",
                color=0xff0000
            )
            await interaction.edit_original_response(embed=embed_erro)
            return

        embed = discord.Embed(
            title="🏰 LOUCOS POR PVE",
            description=f"Informações da guilda",
            color=0x00ff00
        )
        embed.add_field(
            name="📋 Informações Gerais",
            value=f"**Nome:** {guilda_info['name']}\n"
                  f"**Fundador:** {guilda_info['founder']}\n"
                  f"**Membros:** {guilda_info['member_count']}\n"
                  f"**Aliança:** {guilda_info['alliance_tag'] or 'Nenhuma'}",
            inline=False
        )
        kill_fame_formatado = formatar_valor_abreviado(guilda_info['kill_fame'])
        death_fame_formatado = formatar_valor_abreviado(guilda_info['death_fame'])
        embed.add_field(
            name="⚔️ Estatísticas de Combate",
            value=f"**Kill Fame:** {kill_fame_formatado}\n"
                  f"**Death Fame:** {death_fame_formatado}",
            inline=True
        )
        founded_date = guilda_info['founded'][:10]
        embed.add_field(
            name="📅 Fundação",
            value=founded_date,
            inline=True
        )
        embed.set_footer(text="Dados da API oficial do Albion Online")
        await interaction.edit_original_response(embed=embed)
    except Exception as e:
        embed_erro = discord.Embed(
            title="❌ Erro interno",
            description=f"Ocorreu um erro: {str(e)}",
            color=0xff0000
        )
        await interaction.edit_original_response(embed=embed_erro)



# ;
# ;
# ;
# ;
# ;
# ------------------------------------- LISTAR MEMBROS DA GUILDA  ---------------------------------
# ;
# ;
# ;
# ;
# ;

@bot.tree.command(name="tutorial_dg", description="Tutorial passo-a-passo para usar /dg_beneficente (com imagens)")
async def tutorial_dg(interaction: discord.Interaction):
    """Envia um tutorial em 3 partes explicando como usar /dg_beneficente.
    Procura por imagens locais (tutorial_tipos.png, tutorial_nomes.png, tutorial_roles.png)
    e as anexa aos embeds se existirem.
    """
    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar
    await safe_defer(interaction)




    # Parte 1: Selecionar o tipo da DG
    embed1 = discord.Embed(
        title="Tutorial - Parte 1: Selecionar o tipo da DG",
        description=(
            "Escolha o tipo correto da DG ao usar /dg_beneficente:\n\n"
            "• **PONTUAÇÃO**: Para usar essa opção você precisa ter **10 pontos**. Use `/consultar_pontuação @membro` para ver quantos pontos você tem.\n\n"
            "• **SORTEIO**: Apenas quem ganhou um sorteio de DG beneficente na última semana pode usar. Use `/listar_sorteios` para verificar se você está na lista.\n\n"
            "• **PATROCINADOR**: Necessita da TAG de patrocinador. Patrocinadores podem fazer 1 DG por semana. Consulte um Líder ou Braço Direito para se tornar patrocinador.\n\n"
            "• **RECRUTADOR**: Apenas membros com a tag de recrutador podem usar essa opção."
        ),
        color=0x00ff00
    )
    # Anexar imagem se existir
    img1 = "tutorial_tipos.png"
    if os.path.exists(img1):
        file1 = discord.File(img1, filename=img1)
        embed1.set_image(url=f"attachment://{img1}")
        await interaction.followup.send(embed=embed1, file=file1)
    else:
        await interaction.followup.send(embed=embed1)

    # Parte 2: Como digitar os nomes dos participantes
    embed2 = discord.Embed(
        title="Tutorial - Parte 2: Digitar os participantes",
        description=(
            "Ao informar os participantes, marque-os com o @ do Discord para garantir que o bot reconheça corretamente.\n\n"
            "• Use o @mention do jogador (ex: `@Nickname`) em vez de digitar apenas o nome.\n"
            "• Quem está puxando a DG (caller) NÃO precisa se adicionar à lista de participantes — o caller não recebe pontuação.\n"
        ),
        color=0xffa500
    )
    img2 = "tutorial_nomes.png"
    if os.path.exists(img2):
        file2 = discord.File(img2, filename=img2)
        embed2.set_image(url=f"attachment://{img2}")
        await interaction.followup.send(embed=embed2, file=file2)
    else:
        await interaction.followup.send(embed=embed2)

    # Parte 3: Selecionar roles (Tank / Healer)
    embed3 = discord.Embed(
        title="Tutorial - Parte 3: Selecionar as roles corretas",
        description=(
            "Ao finalizar a prévia, selecione apenas o **Tank** e o **Healer**.\n\n"
            "• O restante da party será automaticamente marcado como **DPS**.\n"
            "• Se você for o caller e também for Tank/Healer, só se adicione com essa função — atenção: o caller **não recebe pontuação**.\n\n"
            "Dica: para garantir a distribuição correta, defina claramente o Tank e o Healer antes de finalizar."
        ),
        color=0x0099ff
    )
    img3 = "tutorial_roles.png"
    if os.path.exists(img3):
        file3 = discord.File(img3, filename=img3)
        embed3.set_image(url=f"attachment://{img3}")
        # Mensagem final bonitinha incorporada como footer/field
        embed3.add_field(name="Prontinho!", value=(
            "Prontinhooo, com isso você já tá apto a puxar sua própria beneficente! 🎉\n\n"
            "A média de retorno de cada DG beneficente é de 4-5M para DG T8.1, podendo subir dependendo do nível da DG (T8.2/T8.3).\n\n"
            "Boa sorte meus queridos! — @Klartz"
        ), inline=False)
        await interaction.followup.send(embed=embed3, file=file3)
    else:
        embed3.add_field(name="Prontinho!", value=(
            "Prontinhooo, com isso você já tá apto a puxar sua própria beneficente! 🎉\n\n"
            "A média de retorno de cada DG beneficente é de 4-5M para DG T8.1, podendo subir dependendo do nível da DG (T8.2/T8.3).\n\n"
            "Boa sorte meus queridos! — @Klartz"
        ), inline=False)
        await interaction.followup.send(embed=embed3)


# Comando para zoar um membro com mensagem personalizada
@bot.tree.command(name="zoar", description="Zoar um membro com uma mensagem personalizada")
@app_commands.describe(membro="Membro a ser zoado (menção ou nome)")
async def zoar(interaction: discord.Interaction, membro: discord.Member):

    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar
    
    await safe_defer(interaction)




    # Validar que todos os integrantes fornecidos são mentions no formato <@123> ou <@!123>
    invalid_tokens = []
    mention_pattern = re.compile(r"^<@!?(\d+)>$")
    if membro:
        for parte in membro.split():
            if not mention_pattern.match(parte):
                invalid_tokens.append(parte)

    if invalid_tokens:
        embed_erro = discord.Embed(
            title="❌ Formato inválido - Integrantes",
            description=(
                "Os participantes devem ser informados como mentions do Discord.\n"
                "Por favor, marque cada participante usando `@Nickname` (o bot recebe o formato interno `<@USERID>`).\n\n"
                f"Tokens inválidos: {', '.join(invalid_tokens)}"
            ),
            color=0xff0000
        )
        embed_erro.add_field(name="💡 Como corrigir", value="Use o @ para mencionar cada jogador; ex: `@Klartz`.", inline=False)
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    # Pegar display_name do membro para comparação
    alvo_nome = membro.display_name

    match alvo_nome:
        case "[IPVE] Klartz":
            mensagem = (
                f"{membro.mention} — "
                "Cuidado com esse cara, é um dos espiões mais temidos do albion online, o spy dele é tão bem feito que ele é capaz de se enfiar no meio dos staffs. 🐘🔥\n"
                "E ai quando você menos esperar... virou bolsinha de LOOT!\n"
            )
        case "[PVE]  AnnyCaroline":
            mensagem = (
                f"{membro.mention} — "
                "Ei Anny, já viu um isekai com preludio de morte sendo ataque cardiaco? Porque eu acho que você tá vivendo um agora... "
            )
        case "[PVE] Pedroww284":
            mensagem = (
                f"{membro.mention} — "
                "Loucarada, esse cara é tão louco por pontos que quando ele era filhote se jogou de cabeça no chão só pra ganhar 12"
            )
        case "[PVE] Pedroww284":
            mensagem = (
                f"{membro.mention} — "
                "Loucarada, esse cara é tão louco por pontos que quando ele era filhote se jogou de cabeça no chão só pra ganhar 12"
            )
        case "[IPVE] MatadorDSpam":
            mensagem = (
                f"{membro.mention} — "
                "Levantem suas calças!! Fechem seus zipers!! e GO-GO-GO! o matador de esperma chegou"
            )
        case _:  # 🔧 CASO PADRÃO - funciona para qualquer outro nome
            mensagem = (
                f"{membro.mention} — "
                "Você é irrelevante demais pra tá aqui, volta a jogar Free Fire!"
            )

    # Enviar a mensagem pública no canal onde o comando foi usado
    await interaction.followup.send(mensagem)



# ;
# ;
# ;
# ;
# ;
# ------------------------------------- LISTAR MEMBROS DA GUILDA  ---------------------------------
# ;
# ;
# ;
# ;
# ;


@bot.tree.command(name="membros", description="Lista todos os membros da guilda LOUCOS POR PVE")
async def membros(interaction: discord.Interaction):
    embed_loading = discord.Embed(
        title="🔍 Buscando membros da guilda...",
        description="Consultando API do Albion Online",
        color=0xffa500
    )
    await interaction.response.send_message(embed=embed_loading)

    try:
        guilda_info = await buscar_guilda_por_id(GUILD_ID)
        if not guilda_info:
            embed_erro = discord.Embed(
                title="❌ Erro ao carregar dados",
                description="Não foi possível carregar as informações da guilda",
                color=0xff0000
            )
            await interaction.edit_original_response(embed=embed_erro)
            return  # <-- IMPORTANTE

        membros = await buscar_membros_guilda(guilda_info)
        if not membros:
            embed_erro = discord.Embed(
                title="❌ Erro ao buscar membros",
                description="Não foi possível carregar os membros da guilda. A API pode estar indisponível.",
                color=0xff0000
            )
            await interaction.edit_original_response(embed=embed_erro)
            return  # <-- IMPORTANTE

        # Montar embed com a lista de membros
        embed = discord.Embed(
            title="👥 MEMBROS DA GUILDA",
            description=f"Total: {len(membros)} membros",
            color=0x00ff00
        )

        nomes = [m.get('Name', 'Desconhecido') for m in membros]
        nomes.sort()
        # Discord limita fields a 1024 caracteres, então pode ser necessário dividir em partes
        chunk_size = 40
        for i in range(0, len(nomes), chunk_size):
            chunk = nomes[i:i+chunk_size]
            embed.add_field(
                name=f"Membros {i+1} - {i+len(chunk)}",
                value="\n".join(chunk),
                inline=False
            )

        embed.set_footer(text="Dados da API oficial do Albion Online")
        await interaction.edit_original_response(embed=embed)

    except Exception as e:
        embed_erro = discord.Embed(
            title="❌ Erro interno",
            description=f"Ocorreu um erro: {str(e)}",
            color=0xff0000
        )
        await interaction.edit_original_response(embed=embed_erro)


# ;
# ;
# ;
# ;
# ;
# ------------------------------------- MOSTRAR FAMA DE MEMBRO DA GUILDA ---------------------------------
# ;
# ;
# ;
# ;
# ;

@bot.command()
async def membro(ctx, *, nome_membro):
    """Mostra informações detalhadas de um membro da guilda"""
    
    # Embed de carregamento
    embed_loading = discord.Embed(
        title="🔍 Buscando informações do membro...",
        description=f"Procurando por: **{nome_membro}**",
        color=0xffa500
    )
    loading_msg = await ctx.send(embed=embed_loading)
    
    try:
        # Buscar informações do membro
        membro_info = await buscar_membro_por_nome(nome_membro)
        
        if not membro_info:
            embed_erro = discord.Embed(
                title="❌ Membro não encontrado",
                description=f"Não foi possível encontrar o membro **{nome_membro}** na guilda LOUCOS POR PVE",
                color=0xff0000
            )
            embed_erro.add_field(
                name="💡 Dica",
                value="Verifique se o nome está correto (sem espaços extras) ou use !membros para ver a lista completa",
                inline=False
            )
            await loading_msg.edit(embed=embed_erro)
            return
        
        # Extrair dados do membro
        nome = membro_info.get('Name', 'Desconhecido')
        kill_fame = membro_info.get('KillFame', 0)
        death_fame = membro_info.get('DeathFame', 0)
        fame_ratio = membro_info.get('FameRatio', 0)
        
        # Estatísticas de PvE
        lifetime_stats = membro_info.get('LifetimeStatistics', {})
        pve_stats = lifetime_stats.get('PvE', {})
        pve_total = pve_stats.get('Total', 0)
        
        # Estatísticas de coleta
        gathering_stats = lifetime_stats.get('Gathering', {})
        fiber_total = gathering_stats.get('Fiber', {}).get('Total', 0)
        hide_total = gathering_stats.get('Hide', {}).get('Total', 0)
        ore_total = gathering_stats.get('Ore', {}).get('Total', 0)
        rock_total = gathering_stats.get('Rock', {}).get('Total', 0)
        wood_total = gathering_stats.get('Wood', {}).get('Total', 0)
        gathering_total = gathering_stats.get('All', {}).get('Total', 0)
        
        # Outras estatísticas
        crafting_total = lifetime_stats.get('Crafting', {}).get('Total', 0)
        fishing_fame = lifetime_stats.get('FishingFame', 0)
        farming_fame = lifetime_stats.get('FarmingFame', 0)
        
        # Criar embed com informações do membro
        embed = discord.Embed(
            title=f"👤 {nome}",
            description="📊 **Estatísticas Detalhadas**",
            color=0x00ff00
        )
        
        # Informações de PvP
        embed.add_field(
            name="⚔️ **PvP Stats**",
            value=f"**Kill Fame:** {formatar_valor_abreviado(kill_fame)}\n"
                  f"**Death Fame:** {formatar_valor_abreviado(death_fame)}\n"
                  f"**Fame Ratio:** {fame_ratio:.2f}",
            inline=True
        )
        
        # Informações de PvE
        embed.add_field(
            name="🏰 **PvE Stats**",
            value=f"**Total Fame:** {formatar_valor_abreviado(pve_total)}\n"
                  f"**Crafting:** {formatar_valor_abreviado(crafting_total)}\n"
                  f"**Fishing:** {formatar_valor_abreviado(fishing_fame)}\n"
                  f"**Farming:** {formatar_valor_abreviado(farming_fame)}",
            inline=True
        )
        
        # Adicionar campo vazio para quebra de linha
        embed.add_field(name="\u200b", value="\u200b", inline=True)
        
        # Estatísticas de coleta
        embed.add_field(
            name="🌿 **Gathering Stats**",
            value=f"**Total Gathering:** {formatar_valor_abreviado(gathering_total)}\n"
                  f"🌾 **Fiber:** {formatar_valor_abreviado(fiber_total)}\n"
                  f"🦌 **Hide:** {formatar_valor_abreviado(hide_total)}\n"
                  f"⛏️ **Ore:** {formatar_valor_abreviado(ore_total)}\n"
                  f"🪨 **Rock:** {formatar_valor_abreviado(rock_total)}\n"
                  f"🪵 **Wood:** {formatar_valor_abreviado(wood_total)}",
            inline=True
        )
        
        # Informações da guilda
        embed.add_field(
            name="🏰 **Guild Info**",
            value=f"**Guilda:** {membro_info.get('GuildName', 'N/A')}\n"
                  f"**Aliança:** {membro_info.get('AllianceName', 'Nenhuma')}",
            inline=True
        )

        embed.set_footer(text="Dados da API oficial do Albion Online. Desenvolvido por: @Klartz")

        await loading_msg.edit(embed=embed)
        
    except Exception as e:
        embed_erro = discord.Embed(
            title="❌ Erro interno",
            description=f"Ocorreu um erro: {str(e)}",
            color=0xff0000
        )
        await loading_msg.edit(embed=embed_erro)


# ;
# ;
# ;
# ;
# ;# ------------------------------------- INFORMAÇÕES DO BOT ---------------------------------
# ;
# ;
# ;
# ;
# ;


@bot.command()
async def botinfo(ctx):

    embed = discord.Embed(
        title="🤖 Informações do Bot",
        description="Detalhes sobre o bot de pontuação",
        color=0x00ff00  # Verde
    )

    embed.add_field(name="Nome", value=bot.user.name, inline=True)
    embed.add_field(name="ID", value=bot.user.id, inline=True)
    embed.add_field(name="Criador", value="Lucas (Klartz)", inline=True)
    embed.add_field(name="Comandos Disponíveis", value="!pontuacao, !conteudo, !finalizar, !split, !guilda, !membros, !membro, !botinfo, !comandos", inline=False)
    embed.add_field(name="Versão", value="1.0.0", inline=True)
    embed.set_footer(text="Para mais comandos, use: !comandos. Desenvolvido por: Lucas (Klartz)")
    
    await ctx.send(embed=embed)



# ;
# ;
# ;
# ;
# ;
# ------------------------------------- ADCIONAR PONTOS ---------------------------------
# ;
# ;
# ;
# ;
# ;

@bot.tree.command(name="addpontos", description="Adiciona pontos a um membro")
@app_commands.describe(
    integrantes="Nome do membro ou @mention",
    pontos="Quantidade de pontos para adicionar"
)
async def addpontos(interaction: discord.Interaction, integrantes: str, pontos: int):
    # 🚀 RESPONDER IMEDIATAMENTE
    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar
    
    await safe_defer(interaction)




        # Validar que todos os integrantes fornecidos são mentions no formato <@123> ou <@!123>
    invalid_tokens = []
    mention_pattern = re.compile(r"^<@!?(\d+)>$")
    if integrantes:
        for parte in integrantes.split():
            if not mention_pattern.match(parte):
                invalid_tokens.append(parte)

    if invalid_tokens:
        embed_erro = discord.Embed(
            title="❌ Formato inválido - Integrantes",
            description=(
                "Os participantes devem ser informados como mentions do Discord.\n"
                "Por favor, marque cada participante usando `@Nickname` (o bot recebe o formato interno `<@USERID>`).\n\n"
                f"Tokens inválidos: {', '.join(invalid_tokens)}"
            ),
            color=0xff0000
        )
        embed_erro.add_field(name="💡 Como corrigir", value="Use o @ para mencionar cada jogador; ex: `@Klartz`.", inline=False)
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return
    
    # Verificar se o usuário tem permissão
    if not any(role.name.lower() in ["zelador"] for role in interaction.user.roles):
        embed_erro = discord.Embed(
            title="❌ Sem Permissão",
            description="Apenas zeladores podem gerenciar pontos.",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    try:
        
        membros = []  # CORRIGIDO: garantir que seja lista
        results = []
        if integrantes:
            for parte in integrantes.split():
                nome_limpo = await tratar_mention(interaction, parte)
                membros.append(nome_limpo)

        for item in membros:
            try:
                membro_limpo = await tratar_mention(interaction, item)
                print(f"[DEBUG] Membro original: {item}, Membro limpo: {membro_limpo}")
                nova_pontuacao = adicionar_pontos(membro_limpo, pontos)
                if nova_pontuacao is not None:
                    results.append((membro_limpo, True, nova_pontuacao))
                else:
                    results.append((membro_limpo, False, None))
            except Exception as inner_e:
                results.append((item, False, str(inner_e)))

        # Construir embed resumo
        embed = discord.Embed(title=f"✅ Resultado: adição de pontos", color=0x00ff00)
        embed.set_author(name="Sistema de Pontuação - LOUCOS POR PVE", icon_url=bot.user.avatar.url if bot.user.avatar else None)
        sucesso_lines = []
        erro_lines = []
        for nome, ok, value in results:
            if ok:
                sucesso_lines.append(f"**{nome}** → agora tem **{value}** pts")
            else:
                erro_lines.append(f"**{nome}** → falha ({value})")

        if sucesso_lines:
            embed.add_field(name=f"✅ Sucesso 🔹🔹(+ {pontos})🔹🔹", value="\n".join(sucesso_lines), inline=False)
            embed.set_footer(text="Use /consultar_pontuação <membro> para ver a pontuação atual.")
        if erro_lines:
            embed.add_field(name="❌ Erros", value="\n".join(erro_lines), inline=False)
            embed.set_footer(text="Verifique se os nomes estão corretos e tente novamente.")

        await interaction.followup.send(embed=embed)

    except Exception as e:
        embed_erro = discord.Embed(
            title="❌ Erro",
            description=f"Ocorreu um erro inesperado: {str(e)}",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro)



# ;
# ;
# ;
# ;
# ;
# ------------------------------------- CONSULTAR POONTOS DO MEMBRO ---------------------------------
# ;
# ;
# ;
# ;
# ;


@bot.tree.command(name="consultar_pontuação", description="Consulta a pontuação de um membro")
@app_commands.describe(membro="Nome do membro para consultar")
async def pontos(interaction: discord.Interaction, membro: str):

    if await permitir_comando_apenas_no_canal(interaction, "📋・consultar-pontuação"):
        return  # Se retornar True, não está no canal permitido e o comando deve parar

    await safe_defer(interaction)

    # Validar que todos os integrantes fornecidos são mentions no formato <@123> ou <@!123>
    invalid_tokens = []
    mention_pattern = re.compile(r"^<@!?(\d+)>$")
    if membro:
        for parte in membro.split():
            if not mention_pattern.match(parte):
                invalid_tokens.append(parte)

    if invalid_tokens:
        embed_erro = discord.Embed(
            title="❌ Formato inválido - Integrantes",
            description=(
                "Os participantes devem ser informados como mentions do Discord.\n"
                "Por favor, marque cada participante usando `@Nickname` (o bot recebe o formato interno `<@USERID>`).\n\n"
                f"Tokens inválidos: {', '.join(invalid_tokens)}"
            ),
            color=0xff0000
        )
        embed_erro.add_field(name="💡 Como corrigir", value="Use o @ para mencionar cada jogador; ex: `@Klartz`.", inline=False)
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    membro_limpo = await tratar_mention(interaction, membro)
    pontuacao_atual = obter_pontuacao(membro_limpo)

    embed = discord.Embed(
        title="📊 Consulta de Pontuação",
        color=0x0099ff
    )
    
    nome_exibicao = membro_limpo

    if pontuacao_atual > 0:
        embed.add_field(
            name=f"🏆 {nome_exibicao}",
            value=f"**{pontuacao_atual}** pontos",
            inline=False
        )
    else:
        embed.add_field(
            name=f"😢 {nome_exibicao}😢 ",
            value="Você não possui pontos ainda, por favor participe de DGs beneficentes ou complete atividades para ganhar pontos!",
            inline=False
        )
    await interaction.followup.send(embed=embed)


# ;
# ;
# ;
# ;
# ;
# ------------------------------------- RANKING DE PONTUAÇÃO ---------------------------------
# ;
# ;
# ;
# ;
# ;




@bot.tree.command(name="ranking", description="Mostra o ranking completo de pontuação")
async def ranking(interaction: discord.Interaction):
    ranking_completo = obter_ranking()

    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar


    
    if not ranking_completo:
        embed = discord.Embed(
            title="📊 Ranking de Pontuação",
            description="Nenhum membro possui pontos ainda.",
            color=0xff9900
        )
        await interaction.response.send_message(embed=embed)
        return
    
    embed = discord.Embed(
        title="🏆 RANKING DE PONTUAÇÃO - LOUCOS POR PVE",
        description="Top membros por pontuação",
        color=0xffd700
    )
    
    # Mostrar top 10 (ou todos se menos de 10)
    top_membros = ranking_completo[:10]
    
    ranking_texto = ""
    for i, (nome, pontos) in enumerate(top_membros, 1):
        if i == 1:
            emoji = "🥇"
        elif i == 2:
            emoji = "🥈"
        elif i == 3:
            emoji = "🥉"
        else:
            emoji = f"{i}."
            
        ranking_texto += f"{emoji} **{nome}** - {pontos} pts\n"
    
    embed.add_field(
        name="🏆 Top Membros",
        value=ranking_texto,
        inline=False
    )
    
    if len(ranking_completo) > 10:
        embed.set_footer(text=f"Mostrando top 10 de {len(ranking_completo)} membros")
    else:
        embed.set_footer(text=f"Total: {len(ranking_completo)} membros")
    await atualizar_ranking()
    await interaction.response.send_message(embed=embed)



# ;
# ;
# ;
# ;
# ;
# ------------------------------------- REMOVER PONTOS ---------------------------------
# ;
# ;
# ;
# ;
# ;



@bot.tree.command(name="removerpontos", description="Remove pontos de um membro")
@app_commands.describe(
    integrantes="Nome do membro ou @mention",
    pontos="Quantidade de pontos para remover"
)
async def removerpontos(interaction: discord.Interaction, integrantes: str, pontos: int):
    # 🚀 RESPONDER IMEDIATAMENTE
    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar
    await safe_defer(interaction)




        # Validar que todos os integrantes fornecidos são mentions no formato <@123> ou <@!123>
    invalid_tokens = []
    mention_pattern = re.compile(r"^<@!?(\d+)>$")
    if integrantes:
        for parte in integrantes.split():
            if not mention_pattern.match(parte):
                invalid_tokens.append(parte)

    if invalid_tokens:
        embed_erro = discord.Embed(
            title="❌ Formato inválido - Integrantes",
            description=(
                "Os participantes devem ser informados como mentions do Discord.\n"
                "Por favor, marque cada participante usando `@Nickname` (o bot recebe o formato interno `<@USERID>`).\n\n"
                f"Tokens inválidos: {', '.join(invalid_tokens)}"
            ),
            color=0xff0000
        )
        embed_erro.add_field(name="💡 Como corrigir", value="Use o @ para mencionar cada jogador; ex: `@Klartz`.", inline=False)
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return
    
    # Verificar se o usuário tem permissão
    if not any(role.name.lower() in ["zelador"] for role in interaction.user.roles):
        embed_erro = discord.Embed(
            title="❌ Sem Permissão",
            description="Apenas zeladores podem gerenciar pontos.",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    try:
        
        membros = []  # CORRIGIDO: garantir que seja lista
        results = []
        if integrantes:
            for parte in integrantes.split():
                nome_limpo = await tratar_mention(interaction, parte)
                membros.append(nome_limpo)

        for item in membros:
            try:
                membro_limpo = await tratar_mention(interaction, item)
                print(f"[DEBUG] Membro original: {item}, Membro limpo: {membro_limpo}")
                nova_pontuacao = remover_pontos(membro_limpo, pontos)
                if nova_pontuacao is not None:
                    results.append((membro_limpo, True, nova_pontuacao))
                else:
                    results.append((membro_limpo, False, None))
            except Exception as inner_e:
                results.append((item, False, str(inner_e)))

        # Construir embed resumo
        embed = discord.Embed(title=f"✅ Resultado: subtração de pontos", color=0x00ff00)
        embed.set_author(name="Sistema de Pontuação - LOUCOS POR PVE", icon_url=bot.user.avatar.url if bot.user.avatar else None)
        sucesso_lines = []
        erro_lines = []
        for nome, ok, value in results:
            if ok:
                sucesso_lines.append(f"**{nome}** → agora tem **{value}** pts")
            else:
                erro_lines.append(f"**{nome}** → falha ({value})")

        if sucesso_lines:
            embed.add_field(name=f"✅ Sucesso 🔸🔸(- {pontos})🔸🔸", value="\n".join(sucesso_lines), inline=False)
            embed.set_footer(text="Use /consultar_pontuação <membro> para ver a pontuação atual.")
        if erro_lines:
            embed.add_field(name="❌ Erros", value="\n".join(erro_lines), inline=False)
            embed.set_footer(text="Verifique se os nomes estão corretos e tente novamente.")

        await interaction.followup.send(embed=embed)

    except Exception as e:
        embed_erro = discord.Embed(
            title="❌ Erro",
            description=f"Ocorreu um erro inesperado: {str(e)}",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro)

# ;
# ;
# ;
# ;
# ;
# ------------------------------------- REMOVER MEMBRO ---------------------------------
# ;
# ;
# ;
# ;
# ;



@bot.tree.command(name="remover_membro_pontos", description="Remove membro do sistema de pontos")
@app_commands.describe(
    integrantes="Nome do membro - @mention"
)
async def remover_membro_pontos(interaction: discord.Interaction, integrantes: str):
    # 🚀 RESPONDER IMEDIATAMENTE

    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar
    await safe_defer(interaction)



    
    # Verificar se o usuário tem permissão
    if not any(role.name.lower() in ["zelador"] for role in interaction.user.roles):
        embed_erro = discord.Embed(
            title="❌ Sem Permissão",
            description="Apenas zeladores podem gerenciar pontos.",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    try:
        
        membros = []  # CORRIGIDO: garantir que seja lista
        results = []
        if integrantes:
            nome_limpo = await tratar_mention(interaction, integrantes)
            membros.append(nome_limpo)

        for item in membros:
            try:
                membro_limpo = await tratar_mention(interaction, item)
                print(f"[DEBUG] Membro original: {item}, Membro limpo: {membro_limpo}")
                nova_pontuacao = remover_membro(membro_limpo)
                if nova_pontuacao is not None:
                    results.append((membro_limpo, True, nova_pontuacao))
                else:
                    results.append((membro_limpo, False, None))
            except Exception as inner_e:
                results.append((item, False, str(inner_e)))

        # Construir embed resumo
        embed = discord.Embed(title=f"✅ Resultado: remoção de membro", color=0x00ff00)
        embed.set_author(name="Sistema de Pontuação - LOUCOS POR PVE", icon_url=bot.user.avatar.url if bot.user.avatar else None)
        sucesso_lines = []
        erro_lines = []
        for nome, ok, value in results:
            if ok:
                sucesso_lines.append(f"**{nome}** → foi removido do sistema de pontos")
            else:
                erro_lines.append(f"**{nome}** → falha ({value})")

        if sucesso_lines:
            embed.add_field(name=f"✅ Sucesso", value="\n".join(sucesso_lines), inline=False)
            embed.set_footer(text="Use /consultar_pontuação <membro> para ver a pontuação atual.")
        if erro_lines:
            embed.add_field(name="❌ Erros", value="\n".join(erro_lines), inline=False)
            embed.set_footer(text="Verifique se os nomes estão corretos e tente novamente.")

        await interaction.followup.send(embed=embed)

    except Exception as e:
        embed_erro = discord.Embed(
            title="❌ Erro",
            description=f"Ocorreu um erro inesperado: {str(e)}",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro)


# ;
# ;
# ;
# ;
# ;
# ------------------------------------- ADICIONAR SORTEIO ---------------------------------
# ;
# ;
# ;
# ;
# ;

@bot.tree.command(name="add_sorteio", description="Adiciona uma pessoa à lista de sorteios")
@app_commands.describe(nomes="Nomes das pessoas ou @mention que ganharam o sorteio")
async def add_sorteio(interaction: discord.Interaction, nomes: str):
    # 🚀 RESPONDER IMEDIATAMENTE
    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar
    await safe_defer(interaction)



    # Validar que todos os integrantes fornecidos são mentions no formato <@123> ou <@!123>
    invalid_tokens = []
    mention_pattern = re.compile(r"^<@!?(\d+)>$")
    if nomes:
        for parte in nomes.split():
            if not mention_pattern.match(parte):
                invalid_tokens.append(parte)

    if invalid_tokens:
        embed_erro = discord.Embed(
            title="❌ Formato inválido - Integrantes",
            description=(
                "Os participantes devem ser informados como mentions do Discord.\n"
                "Por favor, marque cada participante usando `@Nickname` (o bot recebe o formato interno `<@USERID>`).\n\n"
                f"Tokens inválidos: {', '.join(invalid_tokens)}"
            ),
            color=0xff0000
        )
        embed_erro.add_field(name="💡 Como corrigir", value="Use o @ para mencionar cada jogador; ex: `@Klartz`.", inline=False)
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return
        
    
    # Verificar se o usuário tem permissão
    if not any(role.name.lower() in ["zelador"] for role in interaction.user.roles):
        embed_erro = discord.Embed(
            title="❌ Sem Permissão",
            description="Apenas zeladores podem gerenciar pontos.",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    try:
        
        membros = []  # CORRIGIDO: garantir que seja lista
        results = []
        if nomes:
            for parte in nomes.split():
                nome_limpo = await tratar_mention(interaction, parte)
                membros.append(nome_limpo)

        for item in membros:
            try:
                membro_limpo = await tratar_mention(interaction, item)
                adicionado = adicionar_sorteio(membro_limpo)
                if adicionado:
                    results.append((membro_limpo, True, None))
                    print(f"Membro adicionado à lista de sorteios: {membro_limpo}")
                else:
                    results.append((membro_limpo, False, None))
            except Exception as inner_e:
                results.append((item, False, str(inner_e)))

        # Construir embed resumo
        embed = discord.Embed(title=f"✅ Resultado: adição ao sorteio 🎲🎲", color=0x00ff00)
        embed.set_author(name="Sistema de Pontuação - LOUCOS POR PVE", icon_url=bot.user.avatar.url if bot.user.avatar else None)
        sucesso_lines = []
        erro_lines = []
        for nome, ok, erro in results:
            if ok:
                sucesso_lines.append(f"**{nome}** → adicionado ao sorteio com sucesso!")
            else:
                erro_lines.append(f"**{nome}** → falha ao adicionar ao sorteio. Erro: {erro}")

        if sucesso_lines:
            embed.add_field(name=f"✅ Sucesso 🎲🎲", value="\n".join(sucesso_lines), inline=False)
            embed.set_footer(text="Use /listar_sorteio para ver a situação atual.")
        if erro_lines:
            embed.add_field(name="❌ Erros", value="\n".join(erro_lines), inline=False)
            embed.set_footer(text="Verifique se os nomes estão corretos e tente novamente.")

        await interaction.followup.send(embed=embed)

    except Exception as e:
        embed_erro = discord.Embed(
            title="❌ Erro",
            description=f"Ocorreu um erro inesperado: {str(e)}",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro)

# ;
# ;
# ;
# ;
# ;
# ------------------------------------- ADICIONAR PATROCINIO ---------------------------------
# ;
# ;
# ;
# ;
# ;




@bot.tree.command(name="add_patrocinio", description="Adiciona uma pessoa à lista de patrocinadores")
@app_commands.describe(nomes="Nomes das pessoas ou @mention que ganharam o patrocinio")
async def add_patrocinio(interaction: discord.Interaction, nomes: str):
    # 🚀 RESPONDER IMEDIATAMENTE
    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar
    await safe_defer(interaction)



    # Validar que todos os integrantes fornecidos são mentions no formato <@123> ou <@!123>
    invalid_tokens = []
    mention_pattern = re.compile(r"^<@!?(\d+)>$")
    if nomes:
        for parte in nomes.split():
            if not mention_pattern.match(parte):
                invalid_tokens.append(parte)

    if invalid_tokens:
        embed_erro = discord.Embed(
            title="❌ Formato inválido - Integrantes",
            description=(
                "Os participantes devem ser informados como mentions do Discord.\n"
                "Por favor, marque cada participante usando `@Nickname` (o bot recebe o formato interno `<@USERID>`).\n\n"
                f"Tokens inválidos: {', '.join(invalid_tokens)}"
            ),
            color=0xff0000
        )
        embed_erro.add_field(name="💡 Como corrigir", value="Use o @ para mencionar cada jogador; ex: `@Klartz`.", inline=False)
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return
        
    
    # Verificar se o usuário tem permissão
    if not any(role.name.lower() in ["zelador"] for role in interaction.user.roles):
        embed_erro = discord.Embed(
            title="❌ Sem Permissão",
            description="Apenas zeladores podem gerenciar pontos.",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    try:
        
        membros = []  # CORRIGIDO: garantir que seja lista
        results = []
        if nomes:
            for parte in nomes.split():
                nome_limpo = await tratar_mention(interaction, parte)
                membros.append(nome_limpo)

        for item in membros:
            try:
                membro_limpo = await tratar_mention(interaction, item)
                adicionado = adicionar_patrocinios(membro_limpo)
                if adicionado:
                    results.append((membro_limpo, True, None))
                    print(f"Membro adicionado à lista de patrocinadores: {membro_limpo}")
                else:
                    results.append((membro_limpo, False, None))
            except Exception as inner_e:
                results.append((item, False, str(inner_e)))

        # Construir embed resumo
        embed = discord.Embed(title=f"✅ Resultado: adição ao patrocinio 💎💎", color=0x00ff00)
        embed.set_author(name="Sistema de Pontuação - LOUCOS POR PVE", icon_url=bot.user.avatar.url if bot.user.avatar else None)
        sucesso_lines = []
        erro_lines = []
        for nome, ok, erro in results:
            if ok:
                sucesso_lines.append(f"**{nome}** → adicionado a lista de patrocinadores com sucesso!")
            else:
                erro_lines.append(f"**{nome}** → falha ao adicionar a lista de patrocinadores. Erro: {erro}")

        if sucesso_lines:
            embed.add_field(name=f"✅ Sucesso 💎💎", value="\n".join(sucesso_lines), inline=False)
            embed.set_footer(text="Use /listar_patrocinadores para ver a situação atual.")
        if erro_lines:
            embed.add_field(name="❌ Erros", value="\n".join(erro_lines), inline=False)
            embed.set_footer(text="Verifique se os nomes estão corretos e tente novamente.")

        await interaction.followup.send(embed=embed)

    except Exception as e:
        embed_erro = discord.Embed(
            title="❌ Erro",
            description=f"Ocorreu um erro inesperado: {str(e)}",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro)


# ;
# ;
# ;
# ;
# ;
# ------------------------------------- REMOVER PATROCINIO ---------------------------------
# ;
# ;
# ;
# ;
# ;

@bot.tree.command(name="remover_patrocinio", description="Remove uma pessoa da lista de patrocinadores")
@app_commands.describe(nomes="Nome da pessoa ou @mention para remover da lista")
async def remover_patrocinio(interaction: discord.Interaction, nomes: str):
    # 🚀 RESPONDER IMEDIATAMENTE
    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar

    await safe_defer(interaction)


    # Validar que todos os integrantes fornecidos são mentions no formato <@123> ou <@!123>
    invalid_tokens = []
    mention_pattern = re.compile(r"^<@!?(\d+)>$")
    if nomes:
        for parte in nomes.split():
            if not mention_pattern.match(parte):
                invalid_tokens.append(parte)

    if invalid_tokens:
        embed_erro = discord.Embed(
            title="❌ Formato inválido - Integrantes",
            description=(
                "Os participantes devem ser informados como mentions do Discord.\n"
                "Por favor, marque cada participante usando `@Nickname` (o bot recebe o formato interno `<@USERID>`).\n\n"
                f"Tokens inválidos: {', '.join(invalid_tokens)}"
            ),
            color=0xff0000
        )
        embed_erro.add_field(name="💡 Como corrigir", value="Use o @ para mencionar cada jogador; ex: `@Klartz`.", inline=False)
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return
    
    # Verificar se o usuário tem permissão
    if not any(role.name.lower() in ["zelador"] for role in interaction.user.roles):
        embed_erro = discord.Embed(
            title="❌ Sem Permissão",
            description="Apenas zeladores podem gerenciar pontos.",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    try:
        
        membros = []  # CORRIGIDO: garantir que seja lista
        results = []
        if nomes:
            for parte in nomes.split():
                nome_limpo = await tratar_mention(interaction, parte)
                membros.append(nome_limpo)

        for item in membros:
            try:
                membro_limpo = await tratar_mention(interaction, item)
                removido = remover_patrocinios(membro_limpo)
                if removido:
                    results.append((membro_limpo, True, None))
                    print(f"Membro removido da lista de patrocinadores: {membro_limpo}")
                else:
                    results.append((membro_limpo, False, None))
            except Exception as inner_e:
                results.append((item, False, str(inner_e)))

        # Construir embed resumo
        embed = discord.Embed(title=f"✅ Resultado: remoção do patrocinador 💎💎", color=0x00ff00)
        embed.set_author(name="Sistema de Pontuação - LOUCOS POR PVE", icon_url=bot.user.avatar.url if bot.user.avatar else None)
        sucesso_lines = []
        erro_lines = []
        for nome, ok, erro in results:
            if ok:
                sucesso_lines.append(f"**{nome}** → removido da lista de patrocinadores com sucesso!")
            else:
                erro_lines.append(f"**{nome}** → falha ao remover da lista de patrocinadores. Erro: {erro}")

        if sucesso_lines:
            embed.add_field(name=f"✅ Sucesso 💎💎", value="\n".join(sucesso_lines), inline=False)
            embed.set_footer(text="Use /listar_patrocinadores para ver a situação atual.")
        if erro_lines:
            embed.add_field(name="❌ Erros", value="\n".join(erro_lines), inline=False)
            embed.set_footer(text="Verifique se os nomes estão corretos e tente novamente.")

        await interaction.followup.send(embed=embed)

    except Exception as e:
        embed_erro = discord.Embed(
            title="❌ Erro",
            description=f"Ocorreu um erro inesperado: {str(e)}",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro)

# ;
# ;
# ;
# ;
# ;
# ------------------------------------- REMOVER SORTEIO ---------------------------------
# ;
# ;
# ;
# ;
# ;

@bot.tree.command(name="remover_sorteio", description="Remove uma pessoa da lista de sorteios")
@app_commands.describe(nomes="Nome da pessoa ou @mention para remover da lista")
async def remover_sorteios(interaction: discord.Interaction, nomes: str):
    # 🚀 RESPONDER IMEDIATAMENTE
    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar
    await safe_defer(interaction)


    
    # Validar que todos os integrantes fornecidos são mentions no formato <@123> ou <@!123>
    invalid_tokens = []
    mention_pattern = re.compile(r"^<@!?(\d+)>$")
    if nomes:
        for parte in nomes.split():
            if not mention_pattern.match(parte):
                invalid_tokens.append(parte)

    if invalid_tokens:
        embed_erro = discord.Embed(
            title="❌ Formato inválido - Integrantes",
            description=(
                "Os participantes devem ser informados como mentions do Discord.\n"
                "Por favor, marque cada participante usando `@Nickname` (o bot recebe o formato interno `<@USERID>`).\n\n"
                f"Tokens inválidos: {', '.join(invalid_tokens)}"
            ),
            color=0xff0000
        )
        embed_erro.add_field(name="💡 Como corrigir", value="Use o @ para mencionar cada jogador; ex: `@Klartz`.", inline=False)
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return
    
    # Verificar se o usuário tem permissão
    if not any(role.name.lower() in ["zelador"] for role in interaction.user.roles):
        embed_erro = discord.Embed(
            title="❌ Sem Permissão",
            description="Apenas zeladores podem gerenciar pontos.",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    try:
        
        membros = []  # CORRIGIDO: garantir que seja lista
        results = []
        if nomes:
            for parte in nomes.split():
                nome_limpo = await tratar_mention(interaction, parte)
                membros.append(nome_limpo)

        for item in membros:
            try:
                membro_limpo = await tratar_mention(interaction, item)
                removido = remover_sorteio(membro_limpo)
                if removido:
                    results.append((membro_limpo, True, None))
                    print(f"Membro removido da lista de sorteios: {membro_limpo}")
                else:
                    results.append((membro_limpo, False, None))
            except Exception as inner_e:
                results.append((item, False, str(inner_e)))

        # Construir embed resumo
        embed = discord.Embed(title=f"✅ Resultado: remoção do sorteio 🎲🎲", color=0x00ff00)
        embed.set_author(name="Sistema de Pontuação - LOUCOS POR PVE", icon_url=bot.user.avatar.url if bot.user.avatar else None)
        sucesso_lines = []
        erro_lines = []
        for nome, ok, erro in results:
            if ok:
                sucesso_lines.append(f"**{nome}** → removido do da lista de sorteados com sucesso!")
            else:
                erro_lines.append(f"**{nome}** → falha ao remover do sorteio. Erro: {erro}")

        if sucesso_lines:
            embed.add_field(name=f"✅ Sucesso 🎲🎲", value="\n".join(sucesso_lines), inline=False)
            embed.set_footer(text="Use /listar_sorteio para ver a situação atual.")
        if erro_lines:
            embed.add_field(name="❌ Erros", value="\n".join(erro_lines), inline=False)
            embed.set_footer(text="Verifique se os nomes estão corretos e tente novamente.")

        await interaction.followup.send(embed=embed)

    except Exception as e:
        embed_erro = discord.Embed(
            title="❌ Erro",
            description=f"Ocorreu um erro inesperado: {str(e)}",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro)



# ;
# ;
# ;
# ;
# ;
# ------------------------------------- LISTAR PATROCINADORES ---------------------------------
# ;
# ;
# ;
# ;
# ;

@bot.tree.command(name="listar_patrocinadores", description="Lista todos os patrocinadores atuais")
async def listar_patrocinadores(interaction: discord.Interaction):

    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar

    patrocinadores = carregar_patrocinadores()

    if not patrocinadores:
        embed = discord.Embed(
            title="📋 Lista de Patrocinadores",
            description="Nenhum patrocinador encontrado.",
            color=0xff9900
        )
        await interaction.response.send_message(embed=embed)
        return
    
    embed = discord.Embed(
        title="💎 Lista de Patrocinadores Ativos",
        description="Patrocinadores que ainda não fizeram a DG da semana e podem puxar:",
        color=0xffd700
    )

    patrocinadores_texto = "\n".join([f"💎 **{nome}**" for nome in patrocinadores])
    embed.add_field(
        name="🏆 Patrocinadores Atuais",
        value=patrocinadores_texto,
        inline=False
    )

    embed.set_footer(text=f"Total: {len(patrocinadores)} pessoas")
    await interaction.response.send_message(embed=embed)
    
# ;
# ;
# ;
# ;
# ;
# ------------------------------------- LISTAR SORTEIO ---------------------------------
# ;
# ;
# ;
# ;
# ;

@bot.tree.command(name="listar_sorteios", description="Lista todas as pessoas que ganharam sorteios")
async def listar_sorteios(interaction: discord.Interaction):

    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar

    sorteios = carregar_sorteios()
    
    if not sorteios:
        embed = discord.Embed(
            title="📋 Lista de Sorteios",
            description="Nenhuma pessoa ganhou sorteios recentemente.",
            color=0xff9900
        )
        await interaction.response.send_message(embed=embed)
        return
    
    embed = discord.Embed(
        title="🎲 Lista de Sorteios Ativos",
        description="Pessoas que ganharam sorteios e podem puxar DGs:",
        color=0xffd700
    )
    
    sorteios_texto = "\n".join([f"🎯 **{nome}**" for nome in sorteios])
    embed.add_field(
        name="🏆 Ganhadores Atuais",
        value=sorteios_texto,
        inline=False
    )
    
    embed.set_footer(text=f"Total: {len(sorteios)} pessoas")
    await interaction.response.send_message(embed=embed)
    

# ;
# ;
# ;
# ;
# ;
# ------------------------------------- LISTAR MEMBROS PONTUACAO ---------------------------------
# ;
# ;
# ;
# ;
# ;

@bot.tree.command(name="listar_pontuacao", description="Mostra a lista completa de todos os membros e seus pontos")
async def listar_pontuacao(interaction: discord.Interaction):

    if await bloquear_comando_no_canal(interaction, "📊・adicionar-pontos-beneficente"):
        return  # Se retornar True, o canal está bloqueado e o comando deve parar

    pontuacao = obter_toda_pontuacao()
    if not pontuacao:
        embed = discord.Embed(
            title="📋 Lista de Pontuação",
            description="Nenhum membro possui pontos ainda.",
            color=0xff9900
        )
        await interaction.response.send_message(embed=embed)
        return

    ranking = sorted(pontuacao.items(), key=lambda x: x[1], reverse=True)
    max_chars = 1000  # Limite seguro por field
    chunk = []
    chunk_len = 0
    embeds = []
    total = len(ranking)
    for idx, (nome, pontos) in enumerate(ranking, 1):
        linha = f"{idx}. **{nome}** - {pontos} pts\n"
        if chunk_len + len(linha) > max_chars and chunk:
            embeds.append((chunk, idx - len(chunk), idx - 1))
            chunk = []
            chunk_len = 0
        chunk.append(linha)
        chunk_len += len(linha)
    if chunk:
        embeds.append((chunk, total - len(chunk) + 1, total))

    # Enviar a primeira resposta
    first_embed = discord.Embed(
        title="🏆 LISTA COMPLETA DE PONTUAÇÃO",
        description=f"Total de membros: {total}",
        color=0x00ff00
    )
    chunk, start, end = embeds[0]
    first_embed.add_field(
        name=f"Membros {start} - {end}",
        value="".join(chunk),
        inline=False
    )
    first_embed.set_footer(text="Use /ranking para ver o top 10.")
    await interaction.response.send_message(embed=first_embed)

    # Enviar o resto como followup
    for chunk, start, end in embeds[1:]:
        embed = discord.Embed(
            title="🏆 LISTA COMPLETA DE PONTUAÇÃO (cont.)",
            description=None,
            color=0x00ff00
        )
        embed.add_field(
            name=f"Membros {start} - {end}",
            value="".join(chunk),
            inline=False
        )
        await interaction.followup.send(embed=embed)


# ;
# ;
# ;
# ;
# ;
# ------------------------------------- Gera arquivo XLSX ---------------------------------
# ;
# ;
# ;
# ;
# ;


# Comando para gerar backup em CSV
@bot.command()
async def backup(ctx):
    """Gera um arquivo xlsx com membros e pontuação."""
    pontuacao = carregar_pontuacao()
    if not pontuacao:
        await ctx.send("Nenhuma pontuação encontrada para backup.")
        return

    caminho_xlsx = "backup_pontuacao.xlsx"
    membros = list(pontuacao.items())
    membros.sort(key=lambda x: x[0].lower())

    # XLSX formatado
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pontuação"
    ws.append(["membros", "pontuação"])
    # Formatação do cabeçalho
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # Adicionar dados
    for membro, pontos in membros:
        ws.append([membro, pontos])
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    # Alinhar colunas
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, max_row=ws.max_row):
        row[0].alignment = Alignment(horizontal="left")
        row[1].alignment = Alignment(horizontal="center")
    wb.save(caminho_xlsx)

    await ctx.send(file=discord.File(caminho_xlsx), content="Backup Excel gerado com sucesso!")



# ;
# ;
# ;
# ;
# ;
# ------------------------------------- CORRIGE NOMES NO DB ---------------------------------
# ;
# ;
# ;
# ;
# ;


# Corrige nomes do DB para o formato com tag do Discord
@bot.command()
async def conserta_db(ctx):
    guild = ctx.guild
    if not guild:
        await ctx.send("Comando só pode ser usado em servidor.")
        return

    role_tag = "Louco por PVE"
    pontuacao = carregar_pontuacao()
    membros_atualizados = 0
    nomes_para_atualizar = []
    
    # async for member in guild.fetch_members(limit=None):
    #     nome_original = member.display_name
    #     nome_limpo = re.sub(r"^\s*\[[^\]]+\]\s*", "", nome_original)
    #     print(f"Nome limpo: {nome_limpo}")

    async for member in guild.fetch_members(limit=None):
        if any(role_tag.lower() in r.name.lower() for r in member.roles):
            # Procura por qualquer nome no JSON que contenha o display_name do membro
            for nome_json in list(pontuacao.keys()):

                nome_original = member.display_name
                nome_limpo = re.sub(r"^\[[^\]]+\]\s*", "", nome_original)
                print(f"Nome limpo: {nome_limpo}")
                # Se o nome no JSON contém o display_name OU começa com uma tag e termina com o display_name
                print(f"[DEBUG] Verificando {nome_json} contra {member.display_name}")

                if nome_json == nome_limpo:
                    if nome_json != member.display_name:
                        pontuacao[member.display_name] = pontuacao.pop(nome_json)
                        membros_atualizados += 1

    salvar_pontuacao(pontuacao)
    await ctx.send(f"Conserto concluído! {membros_atualizados} nomes atualizados no banco de dados.")



# ;
# ;
# ;
# ;
# ;
# ------------------------------------- TROCA DE PONTOS ENTRE MEMBROS ---------------------------------
# ;
# ;
# ;
# ;
# ;



@bot.tree.command(name="troca_de_pontos", description="Transfere pontos para outro membro da guilda")
@app_commands.describe(
    destinatario="Nome do destinatário ou @mention",
    valor="Quantidade de pontos a transferir"
)
async def troca(interaction: discord.Interaction, destinatario: str, valor: int):

    await safe_defer(interaction)

    # Validar que todos os integrantes fornecidos são mentions no formato <@123> ou <@!123>
    invalid_tokens = []
    mention_pattern = re.compile(r"^<@!?(\d+)>$")
    if destinatario:
        for parte in destinatario.split():
            if not mention_pattern.match(parte):
                invalid_tokens.append(parte)

    if invalid_tokens:
        embed_erro = discord.Embed(
            title="❌ Formato inválido - Integrantes",
            description=(
                "Os participantes devem ser informados como mentions do Discord.\n"
                "Por favor, marque cada participante usando `@Nickname` (o bot recebe o formato interno `<@USERID>`).\n\n"
                f"Tokens inválidos: {', '.join(invalid_tokens)}"
            ),
            color=0xff0000
        )
        embed_erro.add_field(name="💡 Como corrigir", value="Use o @ para mencionar cada jogador; ex: `@Klartz`.", inline=False)
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return
    
    remetente_nome = interaction.user.display_name
    # Validar valor
    if valor <= 0:
        await interaction.followup.send("❌ O valor deve ser maior que zero.", ephemeral=True)
        return

    # Tratar mention do destinatário
    destinatario_limpo = await tratar_mention(interaction, destinatario)

    # Conferir se destinatário está na guilda
    membro_destino = discord.utils.find(lambda m: m.display_name.lower() == destinatario_limpo.lower(), interaction.guild.members)
    if not membro_destino:
        await interaction.followup.send(f"❌ O destinatário **{destinatario_limpo}** não foi encontrado na guilda.", ephemeral=True)
        return

    # Conferir se remetente tem saldo suficiente
    saldo_remetente = obter_pontuacao(remetente_nome)
    if saldo_remetente < valor:
        await interaction.followup.send(f"❌ Você não tem pontos suficientes para transferir.\nSeu saldo: {saldo_remetente}", ephemeral=True)
        return

    # Transferir pontos
    adicionar_pontos(remetente_nome, -valor)
    adicionar_pontos(destinatario_limpo, valor)

    embed = discord.Embed(
        title="🔄 Troca de Pontos Realizada",
        description=f"**{remetente_nome}** transferiu **{valor}** pontos para **{destinatario_limpo}**.",
        color=0x00bfff
    )
    embed.add_field(name="Saldo do remetente", value=f"{remetente_nome}: {obter_pontuacao(remetente_nome)} pontos", inline=False)
    embed.add_field(name="Saldo do destinatário", value=f"{destinatario_limpo}: {obter_pontuacao(destinatario_limpo)} pontos", inline=False)
    embed.set_footer(text="dê /consultar_pontuação (@nickname) para ver seu saldo atualizado.")
    await interaction.followup.send(embed=embed)
    # ;
tickets_recrutamento = {}

# ;
# ;
# ;
# ;
# ;
# ----------------------------------------- SISTEMA DE RECRUTAMENTO PARA TICKETS ---------------------------------
# ;
# ;
# ;
# ;
# ;

# @bot.event
# async def on_guild_channel_create(channel):
#     """Monitora a criação de canais e inicia processo de recrutamento para tickets"""
    
#     # Verificar se é um canal de texto
#     if not isinstance(channel, discord.TextChannel):
#         return
    
#     # Verificar se o nome do canal começa com "ticket-"
#     if not channel.name.lower().startswith("ticket-"):
#         return
    
#     print(f"[TICKET] Novo ticket criado: {channel.name} (ID: {channel.id})")
    
#     try:
#         # Aguardar um pouco para garantir que o canal está totalmente criado
#         await asyncio.sleep(2)
        
#         # Inicializar estado do ticket
#         tickets_recrutamento[channel.id] = {
#             "etapa": "boas_vindas",
#             "dados": {}
#         }
        
#         # Enviar mensagem de boas-vindas personalizada
#         mensagem_boas_vindas = (
#             f"🏰 **BEM-VINDO AO PROCESSO DE RECRUTAMENTO - LOUCOS POR PVE!**\n\n"
#             f"Olá! Seja muito bem-vindo(a) ao nosso servidor, você está no ticket: {channel.mention}!\n\n"
#             f"📋 **INFORMAÇÕES IMPORTANTES:**\n"
#             f"• Nossa guild **NÃO aceita menores de 18 anos**\n"
#             f"• Para players com **menos de 30M de fama total**, é necessário ter **indicação de um membro**\n\n"
#             f"❓ **PRIMEIRA PERGUNTA:**\n"
#             f"Você foi **indicado** por algum membro da nossa guild?\n\n"
#             f"🔹 Responda **SIM** se foi indicado por alguém\n"
#             f"🔹 Responda **NÃO** se não foi indicado por ninguém"
#         )
        
#         await channel.send(mensagem_boas_vindas)
        
#         print(f"[TICKET] Processo de recrutamento iniciado para {channel.name}")
        
#     except discord.Forbidden:
#         print(f"[TICKET] Sem permissão para enviar mensagem no canal {channel.name}")
#     except Exception as e:
#         print(f"[TICKET] Erro ao iniciar recrutamento: {e}")

# @bot.event
# async def on_message(message):
#     # Ignorar mensagens do próprio bot
#     if message.author.bot:
#         return

#     # Verificar se é em um canal de ticket
#     if (isinstance(message.channel, discord.TextChannel) and 
#         message.channel.name.lower().startswith("ticket-") and
#         message.channel.id in tickets_recrutamento):
        
#         await processar_etapa_recrutamento(message)
    
#     # Código existente para outros canais
#     if message.channel.name == "🎁・dg-beneficente":
#         # ... código existente ...
#         pass
    
#     await bot.process_commands(message)

# async def processar_etapa_recrutamento(message):
#     """Processa cada etapa do recrutamento"""
#     channel = message.channel
#     user = message.author
#     content = message.content.strip().upper()
    
#     ticket_data = tickets_recrutamento[channel.id]
#     etapa_atual = ticket_data["etapa"]
    
#     print(f"[RECRUTAMENTO] Canal: {channel.name}, Etapa: {etapa_atual}, Mensagem: {content}")

#     aplicou = ticket_data.get("aplicou", False)


#     try:
#         if etapa_atual == "boas_vindas":
#             if content in ["SIM", "S"]:
#                 # Usuário foi indicado - mostrar select com membros
#                 await processar_indicacao_sim(channel, user)
#             elif content in ["NÃO", "NAO", "N"]:
#                 # Usuário não foi indicado - continuar para próxima etapa
#                 await processar_indicacao_nao(channel, user)
#             else:
#                 await channel.send(
#                     f"❌ **Por favor, responda apenas:**\n"
#                     f"• **SIM** - se foi indicado por alguém\n"
#                     f"• **NÃO** - se não foi indicado por ninguém"
#                 )
        
#         elif etapa_atual == "aguardando_indicador":
#             if content == "PRONTO":
#                 await processar_indicacao_nao(channel, user)
        
#         elif etapa_atual == "aguardando_print":
#                 # Verificar se a mensagem tem anexos de imagem
#                 if message.attachments:
#                     # Verificar se pelo menos um anexo é uma imagem
#                     imagens_encontradas = []
#                     for attachment in message.attachments:
#                         if any(attachment.filename.lower().endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.webp']):
#                             imagens_encontradas.append(attachment.filename)
                    
#                     if imagens_encontradas:
#                         # Marcar que a imagem foi enviada
#                         tickets_recrutamento[channel.id]["dados"]["print_enviado"] = True
                        
#                         await channel.send(
#                             f"✅ **Imagem recebida!**\n\n"
#                             f"Print dos status detectado: `{', '.join(imagens_encontradas)}`\n\n"
#                             f"Agora escreva **PRONTO** para continuar para a próxima etapa."
#                         )
#                     else:
#                         await channel.send(
#                             f"❌ **Arquivo não é uma imagem válida!**\n\n"
#                             f"Por favor, envie uma imagem dos seus status (PNG, JPG, GIF, etc.) e depois escreva **PRONTO**."
#                         )
                
#                 elif content == "PRONTO":
#                     # Verificar se já foi enviada uma imagem
#                     print_enviado = ticket_data.get("dados", {}).get("print_enviado", False)
                    
#                     if print_enviado:
#                         await processar_print_enviado(channel, user)
#                     else:
#                         await channel.send(
#                             f"⚠️ **Ops! Você ainda não enviou o print dos seus status!**\n\n"
#                             f"📸 Por favor, primeiro envie uma **imagem/screenshot** dos seus status do Albion Online.\n"
#                             f"Depois disso, escreva **PRONTO** para continuar.\n\n"
#                             f"💡 **Lembre-se:** A imagem deve mostrar suas estatísticas/atributos do personagem no jogo."
#                         )
#                 else:
#                     await channel.send(
#                         f"📸 **Ainda aguardando seu print dos status!**\n\n"
#                         f"• Primeiro: Envie uma **imagem** dos seus status do Albion Online\n"
#                         f"• Depois: Escreva **PRONTO** para continuar\n\n"
#                         f"💡 **Formatos aceitos:** PNG, JPG, JPEG, GIF, WEBP"
#                     )

#         elif etapa_atual == "aguardando_aplicacao":
#             if content == "PRONTO" and aplicou == True:
#                 await processar_aplicacao_feita(channel, user)
#             elif content == "LOUCOS POR PVE":
#                 tickets_recrutamento[channel.id]["aplicou"] = True
#                 await channel.send(
#                     f"Vejo que você já se aplicou para nossa guild \"**LOUCOS POR PVE**\", agora só falta um recrutador aceitar sua aplicação.\n"
#                     f"Caso esteja tudo certo, escreva **PRONTO** para continuar."
#                 )
#             elif content == "INSANOS POR PVE":
#                 tickets_recrutamento[channel.id]["aplicou"] = True
#                 await channel.send(
#                     f"Vejo que você já se aplicou para nossa guild \"**ISANOS POR PVE**\", agora só falta um recrutador aceitar sua aplicação.\n"
#                     f"Caso esteja tudo certo, escreva **PRONTO** para continuar."
#                 )
#             elif content == "FANATICOS POR PVE":
#                 tickets_recrutamento[channel.id]["aplicou"] = True
#                 await channel.send(
#                     f"Vejo que você já se aplicou para nossa guild \"**FANATICOS POR PVE**\", agora só falta um recrutador aceitar sua aplicação.\n"
#                     f"Caso esteja tudo certo, escreva **PRONTO** para continuar."
#                 )
        
#         elif etapa_atual == "aguardando_recrutador":
#             if content == "PRONTO":
#                 # Verificar se quem escreveu "PRONTO" é um recrutador
#                 member = user
#                 if any("recrutador" in role.name.lower() for role in member.roles):
#                     await processar_tutorial_final(channel, user)
#                 else:
#                     await channel.send(
#                         f"😄 **kkkkk boa tentativa!**\n\n"
#                         f"Mas quem precisa te aceitar é um **recrutador**, por favor aguarde enquanto um recrutador aceita sua aplicação :D\n\n"
#                         f"🔍 **Status:** Aguardando aprovação de um membro com TAG de **Recrutador**"
#                     )
    
#     except Exception as e:
#         print(f"[RECRUTAMENTO] Erro ao processar etapa: {e}")

# async def processar_indicacao_sim(channel, user):

#     await channel.send(
#         f"👥 **ÓTIMO! Você foi indicado por alguém.**\n"
#         f"*Por favor, digite quem foi que te indicou *\n"
#         f"`Lembre-se, com menos de 30M de fama, você só vai conseguir aprovação caso tenha sido indicado por alguém`\n\n"
#         f"**Caso esteja tudo pronto digite `PRONTO`**"
#     )
#     tickets_recrutamento[channel.id]["etapa"] = "aguardando_indicador"

# async def processar_indicacao_nao(channel, user):
#     """Usuário disse que não foi indicado - pedir print dos status"""
#     await pedir_print_status(channel, user)

# async def pedir_print_status(channel, user):
#     """Pede para o usuário enviar print dos status do jogo"""
    
#     # Aqui você pode colocar uma imagem de exemplo se tiver
#     exemplo_texto = "*(envie uma imagem similar ao exemplo abaixo)*"  # Substituir por imagem real se tiver
#     img = "tutorial_status.png"
    
#     mensagem_print = (
#         f"📸 **ETAPA 2: PRINT DOS SEUS STATUS**\n\n"
#         f"Agora preciso que você envie um **print/screenshot** dos seus status dentro do Albion Online.\n\n"
#         f"📋 **Como fazer:**\n"
#         f"• Abra o Albion Online\n"
#         f"• Vá na tela de atributos/estatísticas do seu personagem\n"
#         f"• Tire um print/screenshot\n"
#         f"• Envie a imagem aqui no chat\n\n"
#         f"⚠️ **Após enviar a imagem, escreva** `PRONTO` **para continuar!**"
#     )
    
#     # Enviar mensagem de exemplo se a imagem existir
#     if os.path.exists(img):
#         file1 = discord.File(img, filename=img)
#         embed1 = discord.Embed(
#             title="📸 Exemplo de Print dos Status",
#             description=exemplo_texto,
#             color=0x00ff00
#         )
#         embed1.set_image(url=f"attachment://{img}")
    
#     # Enviar a mensagem principal
#     await channel.send(mensagem_print)
#     await channel.send(embed=embed1, file=file1)
#     tickets_recrutamento[channel.id]["etapa"] = "aguardando_print"

# async def processar_print_enviado(channel, user):
#     """Processa quando o print foi enviado"""
    
#     await channel.send(
#         f"✅ **Confirmação recebida!**\n\n"
#         f"Vamos para a próxima etapa!"
#     )
    
#     await asyncio.sleep(2)
#     await pedir_aplicacao_guild(channel, user)

# async def pedir_aplicacao_guild(channel, user):
#     """Pede para o usuário se aplicar em uma das guilds"""
#     gif_tutorial = "tutorial_aplicacao_guild.gif"
    
#     mensagem_aplicacao = (
#         f"🏰 **ETAPA 3: APLICAÇÃO NA GUILD**\n\n"
#         f"`Não temos guild principal todas tem o mesmo nivel de relevância`\n\n"
#         f"Agora você deve se aplicar nas nossas 3 guildas:\n\n"
#         f"🔹 **LOUCOS POR PVE** `Cheio` \n"
#         f"🔹 **INSANOS POR PVE** `Cheio` \n"
#         f"🔹 **FANATICOS POR PVE** `Nova vazio - chance maior` \n\n"
#         f"📋 **Como fazer:**\n"
#         f"• Abra o Albion Online\n"
#         f"• Vá no menu de Guilds\n"
#         f"• Procure por uma das guilds acima\n"
#         f"• Clique em 'Aplicar' ou 'Join'\n\n"
#         f"*Aplique nas 3 para que sua aprovação seja mais rápida*\n\n"
#         f"⚠️ **Após se aplicar, escreva o nome da guild que você escolheu\n\n"
#         f"**Exemplo:** `FANATICOS POR PVE`"
#     )

    
#     # Enviar mensagem principal primeiro
#     await channel.send(mensagem_aplicacao)
    
#     # Enviar o GIF tutorial se ele existir
#     if os.path.exists(gif_tutorial):
#         try:
#             # Criar embed para o GIF
#             embed_gif = discord.Embed(
#                 title="🎮 Tutorial: Como se aplicar na Guild",
#                 description="Siga os passos mostrados no GIF abaixo para se aplicar em uma das nossas guilds",
#                 color=0x00ff00
#             )
            
#             # Anexar o GIF
#             file_gif = discord.File(gif_tutorial, filename=gif_tutorial)
#             embed_gif.set_image(url=f"attachment://{gif_tutorial}")
            
#             await channel.send(embed=embed_gif, file=file_gif)
#             print(f"[TUTORIAL] GIF de aplicação enviado: {gif_tutorial}")
            
#         except Exception as e:
#             print(f"[TUTORIAL] Erro ao enviar GIF: {e}")
#             await channel.send("*(GIF tutorial não disponível no momento)*")
#     else:
#         await channel.send("*(GIF tutorial não encontrado - verifique se o arquivo está na pasta raiz)*")
    
#     tickets_recrutamento[channel.id]["etapa"] = "aguardando_aplicacao"

# async def processar_aplicacao_feita(channel, user):
#     """Processa quando a aplicação foi feita"""
    
#     await channel.send(
#         f"✅ **Aplicação confirmada!**\n\n"
#     )
    
#     await asyncio.sleep(2)
#     await notificar_recrutadores(channel, user)

# async def notificar_recrutadores(channel, user):
#     """Notifica recrutadores sobre o novo candidato"""
    
#     guild = channel.guild
#     recrutadores = []
    
#     # Buscar membros com role "Recrutador"
#     for member in guild.members:
#         if any("recrutador" in role.name.lower() for role in member.roles):
#             recrutadores.append(member)
    
#     if recrutadores:
#         mensagem_notificacao = (
#             f"**-------------------------------------------------------------------------**\n"
#             f"**RECRUTADOR**\n"
#             f"**-------------------------------------------------------------------------**\n\n"
#             f"🔔 **NOVO CANDIDATO PRONTO PARA APROVAÇÃO!**\n\n"
#             f"Ticket: {channel.mention}\n"
#             f"Candidato: {user.mention}\n\n"
#             f"✅ O candidato já completou todas as etapas do processo e está aguardando aprovação na guild.\n\n"
#             f"📋 **Próximos passos:**\n"
#             f"• Revisar aplicação na guild dentro do jogo\n"
#             f"• Aprovar o candidato\n"
#             f"• Após aprovação, escrever `PRONTO` no ticket"
#         )
        
#         # Enviar DM para cada recrutador
#         for recrutador in recrutadores:
#             try:
#                 await recrutador.send(mensagem_notificacao)
#                 print(f"[RECRUTAMENTO] DM enviado para recrutador: {recrutador.display_name}")
#             except discord.Forbidden:
#                 print(f"[RECRUTAMENTO] Não foi possível enviar DM para: {recrutador.display_name}")
        
#         # Também postar no canal do ticket
#         await channel.send(
#             f"📢 **RECRUTADORES NOTIFICADOS!**\n\n"
#             f"Nossos recrutadores foram notificados sobre sua aplicação.\n"
#             f"Aguarde a aprovação dentro do jogo.\n\n"
#             f"⏳ **Recrutador, por favor, escreva** `PRONTO` **para que o novo membro possa ver o tutorial final!**"
#         )
#     else:
#         await channel.send(
#             f"❌ **Erro:** Não foi possível encontrar recrutadores online.\n"
#             f"Por favor, aguarde ou contate um administrador."
#         )
    
#     tickets_recrutamento[channel.id]["etapa"] = "aguardando_recrutador"

# async def processar_tutorial_final(channel, user):
#     """Mostra o tutorial final para o novo membro"""
    
#     # Nome da imagem na pasta raiz
#     img_tutorial = "tutorial_mostrar_canais.png"  # ou o nome que sua imagem tem
    
#     # Primeira parte da mensagem
#     tutorial_inicial = (
#         f"🎉 **PARABÉNS! VOCÊ FOI ACEITO NA GUILD!**\n\n"
#         f"*Para que não tenha nenhum problema com os canais do discord, recomendamos que vá até as configurações do servidor*\n"
#         f"*e habilite* `mostrar todos os canais` como na imagem a seguir:\n"
#     )
    
#     await channel.send(tutorial_inicial)
    
#     # Enviar a imagem se ela existir
#     if os.path.exists(img_tutorial):
#         try:
#             file_tutorial = discord.File(img_tutorial, filename=img_tutorial)
#             embed_img = discord.Embed(
#                 title="📋 Tutorial: Como mostrar todos os canais",
#                 description="Siga os passos mostrados na imagem acima",
#                 color=0x00ff00
#             )
#             embed_img.set_image(url=f"attachment://{img_tutorial}")
#             await channel.send(embed=embed_img, file=file_tutorial)
#         except Exception as e:
#             print(f"[TUTORIAL] Erro ao enviar imagem: {e}")
#             await channel.send("*(Imagem de tutorial não disponível)*")
#     else:
#         await channel.send("*(Imagem de tutorial não encontrada)*")
    
#     # Segunda parte da mensagem
#     tutorial_final = (
#         f"\nBem-vindo(a) oficialmente à família **LOUCOS POR PVE**! {user.mention}\n\n"
#         f"📚 **TUTORIAL FINAL - REGISTRO NO DISCORD:**\n\n"
#         f"Para completar seu processo, você deve se registrar no nosso sistema:\n\n"
#         f"🔹 **Digite o comando:** `/registro`\n"
#         f"🔹 **Quando solicitado, digite seu nickname do jogo Albion Online**\n\n"
#         f"📋 **Exemplo:**\n"
#         f"`/registro` → Digite: `SeuNickDoJogo`\n\n"
#         f"✅ **Após o registro você terá acesso a:**\n"
#         f"• Canais exclusivos da guild\n"
#         f"• Sistema de pontuação\n"
#         f"• DGs beneficentes\n"
#         f"• Eventos e sorteios\n\n"
#         f"🎊 **Mais uma vez, seja muito bem-vindo(a)!**\n"
#         f"Se tiver dúvidas, pode perguntar aqui mesmo ou nos canais da guild."
#     )
    
#     await channel.send(tutorial_final)
    
#     # Limpar dados do ticket
#     if channel.id in tickets_recrutamento:
#         del tickets_recrutamento[channel.id]
    
#     print(f"[RECRUTAMENTO] Processo finalizado para {channel.name}")


# async def processar_problema_com_registro(channel, user):


#     tutorial_final = (
#         f"\nBem-vindo(a) oficialmente à família **LOUCOS POR PVE**! {user.mention}\n\n"
#         f"📚 **TUTORIAL FINAL - REGISTRO NO DISCORD:**\n\n"
#         f"Para completar seu processo, você deve se registrar no nosso sistema:\n\n"
#         f"🔹 **Digite o comando:** `/registro`\n"
#         f"🔹 **Quando solicitado, digite seu nickname do jogo Albion Online**\n\n"
#         f"📋 **Exemplo:**\n"
#         f"`/registro` → Digite: `SeuNickDoJogo`\n\n"
#         f"✅ **Após o registro você terá acesso a:**\n"
#         f"• Canais exclusivos da guild\n"
#         f"• Sistema de pontuação\n"
#         f"• DGs beneficentes\n"
#         f"• Eventos e sorteios\n\n"
#         f"🎊 **Mais uma vez, seja muito bem-vindo(a)!**\n"
#         f"Se tiver dúvidas, pode perguntar aqui mesmo ou nos canais da guild."
#     )
    
#     await channel.send(tutorial_final)
    
#     # Limpar dados do ticket
#     if channel.id in tickets_recrutamento:
#         del tickets_recrutamento[channel.id]
    
#     print(f"[RECRUTAMENTO] Processo finalizado para {channel.name}")

# ;
# ;
# ;
# ;
# ;
# ----------------------------------------- CLASSES PARA SELECT MENU ---------------------------------
# ;
# ;
# ;
# ;
# ;

class IndicadorSelectView(discord.ui.View):
    def __init__(self, opcoes_membros):
        super().__init__(timeout=300)  # 5 minutos
        self.add_item(IndicadorSelectMenu(opcoes_membros))

class IndicadorSelectMenu(discord.ui.Select):
    def __init__(self, opcoes_membros):
        super().__init__(
            placeholder="Selecione quem te indicou...",
            options=opcoes_membros,
            min_values=1,
            max_values=1
        )

    async def callback(self, interaction: discord.Interaction):
        # Buscar o membro selecionado
        member_id = int(self.values[0])
        indicador_member = interaction.guild.get_member(member_id)
        
        if indicador_member:
            await interaction.response.send_message(
                f"✅ **Indicador selecionado:** {indicador_member.display_name}",
                ephemeral=True
            )
            
            # Processar a seleção
            await processar_indicador_selecionado(
                interaction.channel, 
                interaction.user, 
                indicador_member
            )
        else:
            await interaction.response.send_message(
                "❌ Erro ao encontrar o membro selecionado.",
                ephemeral=True
            )

# ;
# ;
# ;
# ;
# ;
# ------------------------------------- REGISTRAR PONTOS COM APROVAÇÃO ---------------------------------
# ;
# ;
# ;
# ;
# ;

# Dicionário para armazenar pedidos de pontos pendentes
pedidos_pontos_pendentes = {}

@bot.tree.command(name="registrar_pontos", description="Solicita registro de pontos para aprovação de um zelador")
@app_commands.describe(
    integrantes="Nomes dos integrantes separados por espaço (@mention)",
    pontos="Quantidade de pontos que cada integrante deve receber"
)
async def registrar_pontos(interaction: discord.Interaction, integrantes: str, pontos: int):
    

    if await permitir_comando_apenas_no_canal(interaction, "💎・solicitar-pontos-cristal"):
        return

    await safe_defer(interaction)

    # Validar que todos os integrantes são mentions válidas
    invalid_tokens = []
    mention_pattern = re.compile(r"^<@!?(\d+)>$")
    if integrantes:
        for parte in integrantes.split():
            if not mention_pattern.match(parte):
                invalid_tokens.append(parte)

    if invalid_tokens:
        embed_erro = discord.Embed(
            title="❌ Formato inválido - Integrantes",
            description=(
                "Os participantes devem ser informados como mentions do Discord.\n"
                "Por favor, marque cada participante usando `@Nickname` (o bot recebe o formato interno `<@USERID>`).\n\n"
                f"Tokens inválidos: {', '.join(invalid_tokens)}"
            ),
            color=0xff0000
        )
        embed_erro.add_field(name="💡 Como corrigir", value="Use o @ para mencionar cada jogador; ex: `@Klartz`.", inline=False)
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    # Validar pontos
    if pontos <= 0:
        embed_erro = discord.Embed(
            title="❌ Valor Inválido",
            description="A quantidade de pontos deve ser maior que zero.",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    # Converter mentions para nomes
    membros_limpos = []

    membros_limpos.append(interaction.user.display_name)


   # Depois adicionar os outros integrantes
    for parte in integrantes.split():
        nome_limpo = await tratar_mention(interaction, parte)
        # Evitar duplicatas (caso a pessoa se mencione)
        if nome_limpo not in membros_limpos:
            membros_limpos.append(nome_limpo)

    # Verificar se o usuário já tem um pedido pendente
    user_id = interaction.user.id
    if user_id in pedidos_pontos_pendentes:
        embed_erro = discord.Embed(
            title="❌ Pedido em Andamento",
            description="Você já tem um pedido de pontos aguardando aprovação. Aguarde a resposta do zelador.",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed_erro, ephemeral=True)
        return

    # Armazenar dados do pedido
    pedidos_pontos_pendentes[user_id] = {
        "solicitante": interaction.user.display_name,
        "solicitante_mention": interaction.user.mention,
        "membros": membros_limpos,
        "pontos_cada": pontos,
        "timestamp": datetime.datetime.now()
    }

    # Criar embed do pedido
    embed_pedido = discord.Embed(
        title="📋 SOLICITAÇÃO DE REGISTRO DE PONTOS",
        description=f"**Solicitante:** {interaction.user.mention}\n**Aguardando aprovação de um Zelador**\n\n",
        color=0xffa500
    )

    # Lista de integrantes
    lista_membros = "\n".join([f"• **{membro}**" for membro in membros_limpos])
    embed_pedido.add_field(
        name="**---- Integrantes 👥 ---- **\n\n",
        value=lista_membros,
        inline=False
    )

    # Pontos por pessoa
    embed_pedido.add_field(
        name="**---- Pontuação 📊 ----**\n",
        value=f"\n\n**{pontos}** pontos para cada integrante\n",
        inline=False
    )

    embed_pedido.add_field(
        name="⚠️ Importante",
        value="Apenas membros com TAG de **Zelador** podem aprovar ou recusar este pedido.",
        inline=False
    )

    embed_pedido.set_footer(text="Use os botões abaixo para aprovar ou recusar")

    # Criar view com botões
    view = AprovacaoPontosView(user_id)
    
    message = await interaction.followup.send(embed=embed_pedido, view=view)
    
    # Salvar referência da mensagem para poder editá-la depois
    pedidos_pontos_pendentes[user_id]["message"] = message

class AprovacaoPontosView(discord.ui.View):
    def __init__(self, solicitante_id):
        super().__init__(timeout=604800)  # 1 semana
        self.solicitante_id = solicitante_id

    @discord.ui.button(label="✅ Aceitar", style=discord.ButtonStyle.success)
    async def aceitar_pedido(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.processar_decisao(interaction, True)

    @discord.ui.button(label="❌ Recusar", style=discord.ButtonStyle.secondary)
    async def recusar_pedido(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.processar_decisao(interaction, False)

    async def processar_decisao(self, interaction: discord.Interaction, aprovado: bool):
        # Verificar se tem permissão de zelador
        if not any("zelador" in role.name.lower() for role in interaction.user.roles):
            await interaction.response.send_message(
                "❌ **Sem Permissão**\n\nApenas membros com TAG de **Zelador** podem aprovar ou recusar pedidos de pontos.",
                ephemeral=True
            )
            return

        # Verificar se o pedido ainda existe
        if self.solicitante_id not in pedidos_pontos_pendentes:
            await interaction.response.send_message(
                "❌ Este pedido não foi encontrado ou já foi processado.",
                ephemeral=True
            )
            return

        # Buscar dados do pedido
        pedido_data = pedidos_pontos_pendentes[self.solicitante_id]
        
        if aprovado:
            await self.processar_aprovacao(interaction, pedido_data)
        else:
            await self.processar_recusa(interaction, pedido_data)

        # Limpar dados do pedido
        del pedidos_pontos_pendentes[self.solicitante_id]

    async def processar_aprovacao(self, interaction: discord.Interaction, pedido_data):
        """Processa a aprovação do pedido"""
        
        # Adicionar pontos para cada membro
        resultados = []
        for membro in pedido_data["membros"]:
            nova_pontuacao = adicionar_pontos(membro, pedido_data["pontos_cada"])
            if nova_pontuacao is not None:
                resultados.append(f"• **{membro}** → +{pedido_data['pontos_cada']} pts (Total: {nova_pontuacao})")
            else:
                resultados.append(f"• **{membro}** → ❌ Erro ao adicionar pontos")

        # Criar embed de aprovação
        embed_aprovado = discord.Embed(
            title="✅ PEDIDO APROVADO",
            description=f"**Aprovado por:** {interaction.user.mention}\n**Solicitante:** {pedido_data['solicitante_mention']}",
            color=0x00ff00
        )

        embed_aprovado.add_field(
            name="📊 Pontos Registrados",
            value="\n".join(resultados),
            inline=False
        )

        embed_aprovado.add_field(
            name="📋 Resumo",
            value=f"**Pontos por pessoa:** {pedido_data['pontos_cada']}\n**Total de integrantes:** {len(pedido_data['membros'])}\n**Total de pontos distribuídos:** {pedido_data['pontos_cada'] * len(pedido_data['membros'])}",
            inline=False
        )

        embed_aprovado.set_footer(text="Pontos registrados no sistema com sucesso! 🎉")

        # Desabilitar botões
        for item in self.children:
            item.disabled = True

        await interaction.response.edit_message(embed=embed_aprovado, view=self)

        # Enviar notificação para o solicitante
        try:
            solicitante = interaction.guild.get_member(self.solicitante_id)
            if solicitante:
                embed_notificacao = discord.Embed(
                    title="✅ Seu pedido foi APROVADO!",
                    description=f"Seus pontos foram registrados no sistema por {interaction.user.mention}",
                    color=0x00ff00
                )
                embed_notificacao.add_field(
                    name="📊 Detalhes",
                    value=f"**{pedido_data['pontos_cada']}** pontos para cada um dos **{len(pedido_data['membros'])}** integrantes",
                    inline=False
                )
                await solicitante.send(embed=embed_notificacao)
        except:
            pass  # Falha silenciosa se não conseguir enviar DM

    async def processar_recusa(self, interaction: discord.Interaction, pedido_data):
        """Processa a recusa do pedido"""
        
        # Criar embed de recusa
        embed_recusado = discord.Embed(
            title="❌ PEDIDO RECUSADO",
            description=f"**Recusado por:** {interaction.user.mention}\n**Solicitante:** {pedido_data['solicitante_mention']}",
            color=0xff0000
        )

        embed_recusado.add_field(
            name="📋 Pedido Original",
            value=f"**Integrantes:** {', '.join(pedido_data['membros'])}\n**Pontos cada:** {pedido_data['pontos_cada']}",
            inline=False
        )

        embed_recusado.add_field(
            name="💡 Orientações para o Solicitante",
            value=(
                "• Revise as regras de pontuação da guild\n"
                "• Verifique se a atividade realizada realmente merece pontos\n"
                "• Certifique-se de que todos os participantes são membros ativos\n"
                "• Quando estiver tudo correto, faça um novo pedido"
            ),
            inline=False
        )

        embed_recusado.set_footer(text="Pedido rejeitado. Revise as informações e tente novamente.")

        # Desabilitar botões
        for item in self.children:
            item.disabled = True

        await interaction.response.edit_message(embed=embed_recusado, view=self)

        # Enviar notificação para o solicitante
        try:
            solicitante = interaction.guild.get_member(self.solicitante_id)
            if solicitante:
                embed_notificacao = discord.Embed(
                    title="❌ Seu pedido foi RECUSADO",
                    description=f"Seu pedido de pontos foi recusado por {interaction.user.mention}",
                    color=0xff0000
                )
                embed_notificacao.add_field(
                    name="🔄 Próximos Passos",
                    value=(
                        "1. **Revise as regras** de pontuação da guild\n"
                        "2. **Verifique** se a atividade merece pontuação\n"
                        "3. **Certifique-se** de que todos os dados estão corretos\n"
                        "4. **Faça um novo pedido** quando estiver tudo certo\n\n"
                        "Use `/registrar_pontos` novamente quando necessário."
                    ),
                    inline=False
                )
                await solicitante.send(embed=embed_notificacao)
        except:
            pass  # Falha silenciosa se não conseguir enviar DM

    async def on_timeout(self):
        """Processa quando o tempo limite expira"""
        
        if self.solicitante_id in pedidos_pontos_pendentes:
            pedido_data = pedidos_pontos_pendentes[self.solicitante_id]
            
            # Criar embed de timeout
            embed_timeout = discord.Embed(
                title="⏰ PEDIDO EXPIRADO",
                description=f"**Solicitante:** {pedido_data['solicitante_mention']}\n\nO pedido expirou por falta de resposta dos zeladores.",
                color=0x888888
            )
            
            embed_timeout.add_field(
                name="🔄 Como proceder",
                value="Faça um novo pedido usando `/registrar_pontos` se ainda precisar registrar esses pontos.",
                inline=False
            )

            # Desabilitar botões
            for item in self.children:
                item.disabled = True

            # Tentar editar a mensagem
            try:
                message = pedido_data.get("message")
                if message:
                    await message.edit(embed=embed_timeout, view=self)
            except:
                pass

            # Limpar dados do pedido
            del pedidos_pontos_pendentes[self.solicitante_id]


@bot.tree.command(name="limpar_pedidos_antigos", description="Remove pedidos de pontos com mais de 7 dias")
async def limpar_pedidos_antigos(interaction: discord.Interaction):
    if not any("zelador" in role.name.lower() for role in interaction.user.roles):
        await interaction.response.send_message("❌ Apenas zeladores podem usar este comando.", ephemeral=True)
        return
    
    agora = datetime.datetime.now()
    pedidos_removidos = 0
    
    for user_id, dados in list(pedidos_pontos_pendentes.items()):
        timestamp = dados.get("timestamp")
        if timestamp and (agora - timestamp).days > 7:  # Mais de 7 dias
            del pedidos_pontos_pendentes[user_id]
            pedidos_removidos += 1
    
    await interaction.response.send_message(f"✅ {pedidos_removidos} pedidos antigos foram removidos.", ephemeral=True)


# ;
# ;
# ;
# ;
# ;
# ;
# ;
# ;
# ;
# ;
# -------------------------------------- FUNÇÕES AUXILIARES ---------------------------------
# ;
# ;
# ;
# ;
# ;
# ;
# ;
# ;
# ;
# ;
# ;
# ;
# ;


async def tratar_mention(interaction: discord.Interaction, nome_ou_mention: str):
    """Converte mention para nome real, se necessário"""
    nome_limpo = nome_ou_mention
    
    if nome_ou_mention.startswith("<@") and nome_ou_mention.endswith(">"):
        user_id = nome_ou_mention.replace("<@", "").replace("!", "").replace(">", "")
        try:
            user = interaction.guild.get_member(int(user_id))
            if not user:
                user = await interaction.guild.fetch_member(int(user_id))
            if user:
                nome_limpo = user.display_name
        except Exception:
            # Se não conseguir converter o mention, manter o nome original
            pass
    
    return nome_limpo

def carregar_sorteios():
    """Carrega a lista de pessoas que ganharam sorteios"""
    if os.path.exists(ARQUIVO_SORTEIOS):
        try:
            with open(ARQUIVO_SORTEIOS, 'r', encoding='utf-8') as arquivo:
                return json.load(arquivo)
        except (json.JSONDecodeError, FileNotFoundError):
            print("❌ Erro ao carregar arquivo de sorteios. Criando novo arquivo...")
            return []
    return []

def carregar_patrocinadores():
    """Carrega a lista de patrocinadores do arquivo JSON"""
    if os.path.exists(ARQUIVO_PATROCINADOR):
        try:
            with open(ARQUIVO_PATROCINADOR, 'r', encoding='utf-8') as arquivo:
                return json.load(arquivo)
        except (json.JSONDecodeError, FileNotFoundError):
            print("❌ Erro ao carregar arquivo de patrocinadores. Criando novo arquivo...")
            return []
    return []

def salvar_sorteios(lista_sorteios):
    """Salva a lista de sorteios no arquivo JSON"""
    try:
        with open(ARQUIVO_SORTEIOS, 'w', encoding='utf-8') as arquivo:
            json.dump(lista_sorteios, arquivo, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"❌ Erro ao salvar sorteios: {e}")
        return False

def salvar_patrocinadores(lista_patrocinadores):
    """Salva a lista de patrocinadores no arquivo JSON"""
    try:
        with open(ARQUIVO_PATROCINADOR, 'w', encoding='utf-8') as arquivo:
            json.dump(lista_patrocinadores, arquivo, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"❌ Erro ao salvar patrocinadores: {e}")
        return False
    
def verificar_tag_discord(member, tag_nome):
    """Verifica se o membro tem uma TAG específica no Discord"""
    if not member or not member.roles:
        return False
    
    # Procurar pela role/tag específica (case insensitive)
    for role in member.roles:
        if tag_nome.lower() in role.name.lower():
            return True
    return False

def adicionar_sorteio(nome_pessoa):
    """Adiciona uma pessoa à lista de sorteios"""
    sorteios = carregar_sorteios()
    if nome_pessoa not in sorteios:
        sorteios.append(nome_pessoa)
        salvar_sorteios(sorteios)
    return True

def remover_sorteio(nome_pessoa):
    """Remove uma pessoa da lista de sorteios"""
    sorteios = carregar_sorteios()
    if nome_pessoa in sorteios:
        sorteios.remove(nome_pessoa)
        salvar_sorteios(sorteios)
        return True
    return False

def adicionar_patrocinios(nome_pessoa):
    """Adiciona uma pessoa à lista de patrocinadores"""
    patrocinadores = carregar_patrocinadores()
    if nome_pessoa not in patrocinadores:
        patrocinadores.append(nome_pessoa)
        salvar_patrocinadores(patrocinadores)
    return True

def remover_patrocinios(nome_pessoa):
    """Remove uma pessoa da lista de patrocinadores"""
    patrocinadores = carregar_patrocinadores()
    if nome_pessoa in patrocinadores:
        patrocinadores.remove(nome_pessoa)
        salvar_patrocinadores(patrocinadores)
        return True
    return False

def verificar_sorteio(nome_pessoa):
    """Verifica se a pessoa ganhou algum sorteio"""
    sorteios = carregar_sorteios()
    return nome_pessoa in sorteios

def carregar_pontuacao():
    """Carrega a pontuação dos membros do arquivo JSON"""
    if os.path.exists(ARQUIVO_PONTUACAO):
        try:
            with open(ARQUIVO_PONTUACAO, 'r', encoding='utf-8') as arquivo:
                return json.load(arquivo)
        except (json.JSONDecodeError, FileNotFoundError):
            print("❌ Erro ao carregar arquivo de pontuação. Criando novo arquivo...")
            return {}
    return {}

def salvar_pontuacao(pontuacao_dict):
    """Salva a pontuação dos membros no arquivo JSON"""
    try:
        with open(ARQUIVO_PONTUACAO, 'w', encoding='utf-8') as arquivo:
            json.dump(pontuacao_dict, arquivo, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"❌ Erro ao salvar pontuação: {e}")
        return False

def adicionar_pontos(nome_membro, pontos):
    """Adiciona pontos a um membro específico"""
    pontuacao = carregar_pontuacao()
    
    # Se o membro já existe, soma os pontos; senão, cria entrada nova
    if nome_membro in pontuacao:
        pontuacao[nome_membro] += pontos
    else:
        if pontos > 0:
            print(f"Adicionando novo membro à pontuação: {nome_membro} com {pontos} pontos")
            pontuacao[nome_membro] = pontos
        else:
            print(f"Tentativa de criar membro {nome_membro} com pontos negativos ou zero ignorada.")
            return None
    # Salvar no arquivo
    if salvar_pontuacao(pontuacao):
        return pontuacao[nome_membro]  # Retorna a pontuação total atual
    return None

def remover_pontos(nome_membro, pontos):
    """Remove pontos de um membro específico"""
    pontuacao = carregar_pontuacao()
    
    # Se o membro já existe, subtrai os pontos; senão, cria entrada nova
    if nome_membro in pontuacao:
        pontuacao[nome_membro] -= pontos
    # Salvar no arquivo
    if salvar_pontuacao(pontuacao):
        return pontuacao[nome_membro]  # Retorna a pontuação total atual
    return None

def obter_pontuacao(nome_membro):
    """Obtém a pontuação atual de um membro específico"""
    pontuacao = carregar_pontuacao()
    return pontuacao.get(nome_membro, 0)

def obter_ranking():
    """Obtém o ranking completo ordenado por pontuação"""
    pontuacao = carregar_pontuacao()
    # Ordenar por pontuação (maior para menor)
    ranking = sorted(pontuacao.items(), key=lambda x: x[1], reverse=True)
    return ranking

def remover_membro(nome_membro):
    """Remove um membro da lista de pontuação"""
    pontuacao = carregar_pontuacao()
    if nome_membro in pontuacao:
        del pontuacao[nome_membro]
        salvar_pontuacao(pontuacao)
        return True
    return False

def resetar_pontuacao():
    """Reseta toda a pontuação (cuidado!)"""
    return salvar_pontuacao({})

# Função para retornar toda a lista de pontuação dos membros
def obter_toda_pontuacao():
    """Retorna o dicionário completo de pontuação dos membros."""
    return carregar_pontuacao()


# Função para atualizar patrocinadores
async def atualizar_patrocinadores():
    print("[TAREFA] Atualizando lista de patrocinadores...")
    for guild in bot.guilds:
        patrocinadores = []
        async for member in guild.fetch_members(limit=None):
            if any('patrocinador' in r.name.lower() for r in member.roles):
                # Nome limpo sem tag
                nome_original = member.display_name
                patrocinadores.append(nome_original)
        # Salva no JSON
        try:
            async with aiofiles.open(ARQUIVO_PATROCINADOR, 'w', encoding='utf-8') as f:
                await f.write(json.dumps(patrocinadores, ensure_ascii=False, indent=2))
            print(f"[TAREFA] {len(patrocinadores)} patrocinadores salvos em {ARQUIVO_PATROCINADOR}")
        except Exception as e:
            print(f"[TAREFA] Erro ao salvar patrocinadores: {e}")

# Função de agendamento semanal
async def agendar_atualizacao_patrocinadores():
    tz = pytz.timezone('America/Sao_Paulo')
    while True:
        now = datetime.datetime.now(tz)

        # Define próxima execução para hoje às 16:00, ou amanhã se já passou
        proxima_execucao = now.replace(hour=0, minute=0, second=0, microsecond=0)
        if proxima_execucao <= now:
            proxima_execucao += datetime.timedelta(days=1)

        segundos_ate_execucao = (proxima_execucao - now).total_seconds()
        horas, resto = divmod(int(segundos_ate_execucao), 3600)
        minutos, _ = divmod(resto, 60)
        tempo_str = f"{horas}h {minutos}min" if horas > 0 else f"{minutos}min"

        print(f"[TAREFA] Próxima notificação de patrocinadores agendada para {proxima_execucao} (em {segundos_ate_execucao:.0f}s)")

        # Espera até 00:00
        await asyncio.sleep(segundos_ate_execucao)

        # Ao acordar, se for domingo (weekday==6), resetar a lista de patrocinadores
        agora = datetime.datetime.now(tz)
        try:
            if agora.weekday() == 6:  # Sunday
                print("[TAREFA] Domingo detectado — recarregando lista de patrocinadores (reset semanal)")
                try:
                    await atualizar_patrocinadores()
                except Exception as e:
                    print(f"[TAREFA] Erro ao resetar patrocinadores: {e}")

            # Carregar quem ainda não fez a DG (lista atualizada)
            patrocinadores_pendentes = carregar_patrocinadores()
        except Exception as e:
            print(f"[TAREFA] Erro ao carregar patrocinadores pendentes: {e}")
            patrocinadores_pendentes = []

        if patrocinadores_pendentes:
            lista_pendentes = '\n'.join(f"• {nome}" for nome in patrocinadores_pendentes)
            texto_pendentes = f"\n\n**Patrocinadores que ainda não fizeram a DG dessa semana:**\n{lista_pendentes}"
            texto_pendentes += "\n\n*Lembrete: Se não fizer a DG até o reset semanal (domingo 00:00), perderá a chance desta semana!*"
        else:
            texto_pendentes = "\n\nTodos os patrocinadores já fizeram a DG beneficente nesta semana!"

        mensagem = (
            f"📢 **Atenção Patrocinadores!**\n\n"
            f"Lista atualizada: **{agora.strftime('%d/%m %H:%M')}**\n"
            f"Se não fizer a DG beneficente até o reset semanal (domingo 00:00), perderá a chance de participar desta semana."
            f"{texto_pendentes}"
        )

        for guild in bot.guilds:
            canal = discord.utils.get(guild.text_channels, name="📢🔸comunicados")
            if canal:
                try:
                    await canal.send(mensagem)
                except Exception as e:
                    print(f"[TAREFA] Erro ao enviar mensagem no canal de comunicados: {e}")

        # Após enviar a notificação, não atualizar imediatamente (o reset só ocorre no domingo)

class FuncoesEquipeView(discord.ui.View):
    def __init__(self, membros, interaction_user, user_id):
        super().__init__(timeout=120)
        self.membros = membros
        self.roles = {m: "DPS" for m in membros}
        self.interaction_user = interaction_user
        self.user_id = user_id
        self.tank_set = False
        self.healer_set = False
        self.interaction = None
        self.mensagens_ephemeral = {}

        for membro in membros:
            self.add_item(FuncoesEquipeButton(membro, self))
        
        self.add_item(FinalizarButton())

    async def on_timeout(self):
        """Limpar dados quando a view expira"""
        if self.user_id in conteudos_em_aberto:
            del conteudos_em_aberto[self.user_id]
        
        # Desabilitar todos os botões
        for item in self.children:
            item.disabled = True
        
        if hasattr(self, "message") and self.message:
            embed_timeout = discord.Embed(
                title="⏰ Comando Expirado",
                description="O comando expirou por inatividade. Use `/dg_beneficente` novamente.",
                color=0x888888
            )
            try:
                await self.message.edit(embed=embed_timeout, view=self)
            except Exception:
                pass

    async def update_embed(self, interaction):
        """Atualizar o embed com as funções atuais"""
        # Buscar dados específicos deste usuário
        user_id = self.user_id
        if user_id not in conteudos_em_aberto:
            try:
                await interaction.response.send_message(
                    "❌ Dados do comando não encontrados.", ephemeral=True
                )
            except:
                pass
            return

        conteudo_dados = conteudos_em_aberto[user_id]
        tipo_valor = conteudo_dados["tipo"]
        
        # Buscar ícone e nome formatado do tipo
        icone = icones.get(tipo_valor, "📋")
        
        embed = discord.Embed(
            title=f"📊 PRÉVIA DE PONTUAÇÃO",
            description=f"{icone} **{tipo_valor}**\n\nClique nos botões abaixo para definir Tank e Healer.\nQuando terminar, clique em **Finalizar**.",
            color=0xffa500
        )
        
        # Limpar fields existentes
        embed.clear_fields()
        
        # Adicionar cada membro com sua função atual
        for membro in self.membros:
            funcao = self.roles[membro]
            if funcao == "TANK":
                emoji = "🛡️"
            elif funcao == "HEALER":
                emoji = "💚"
            else:
                emoji = "⚔️"
            
            embed.add_field(
                name=f"{emoji} {membro}",
                value=f"Função: **{funcao}**",
                inline=False
            )
        
        # Atualizar a mensagem original e responder ao select
        try:
            await self.message.edit(embed=embed, view=self)
            # Responder ao select menu
            try:
                await interaction.response.send_message(
                    f"✅ Função atualizada com sucesso!", 
                    ephemeral=True
                )
            except:
                try:
                    await interaction.followup.send(
                        f"✅ Função atualizada com sucesso!", 
                        ephemeral=True
                    )
                except:
                    pass
        except Exception as e:
            print(f"[ERROR] Erro ao atualizar embed: {e}")
            try:
                await interaction.response.send_message(
                    "✅ Função atualizada!", ephemeral=True
                )
            except:
                try:
                    await interaction.followup.send(
                        "✅ Função atualizada!", ephemeral=True
                    )
                except:
                    pass


class FinalizarButton(discord.ui.Button):
    def __init__(self):
        super().__init__(label="✅ Finalizar", style=discord.ButtonStyle.success, row=4)

    async def callback(self, interaction: discord.Interaction):
        global conteudos_em_aberto  # Adicionar esta linha
        resumo_texto = ""
        
        if interaction.user != self.view.interaction_user:
            await interaction.response.send_message(
                "❌ Apenas quem usou o comando pode finalizar.", ephemeral=True
            )
            return

        # Buscar dados específicos deste usuário
        user_id = self.view.user_id
        if user_id not in conteudos_em_aberto:
            await interaction.response.send_message(
                "❌ Dados do comando não encontrados. Tente usar `/dg_beneficente` novamente.", 
                ephemeral=True
            )
            return

        conteudo_em_aberto_dados = conteudos_em_aberto[user_id]  # Corrigir esta linha

        # Calcular pontuação
        pontuacao = {}
        caller_nome = conteudo_em_aberto_dados["caller"]  # Usar a variável corrigida
        tipo_conteudo = conteudo_em_aberto_dados["tipo"]   # Usar a variável corrigida
        membros_sem_caller = [m for m in self.view.membros if m != caller_nome]

        for membro in membros_sem_caller:
            funcao = self.view.roles[membro]
            if funcao in ["TANK", "HEALER"]:
                pontos = 2
            else:
                pontos = 1
            pontuacao[membro] = {"funcao": funcao, "pontos": pontos}

        # NOVO: Verificar se o tipo é "PONTUAÇÃO" e penalizar o caller
        if tipo_conteudo == "PONTUAÇÃO":
            remover_pontos(caller_nome, 10)
            penalidade_texto = f"**Caller: ** 👑 {caller_nome}  ⛔10 pontos (PONTUAÇÃO)"
        else:
            penalidade_texto = f"**Caller: ** 👑 {caller_nome} (não recebe pontos)"

        if tipo_conteudo == "SORTEIO":
            remover_sorteio(caller_nome)
            penalidade_texto += " | Sorteio removido da lista. você não pode ser soteado novamente em 3 dias."
        if tipo_conteudo == "PATROCIONADOR":
            remover_patrocinios(caller_nome)
            penalidade_texto += " | Patrocínio removido da lista. recarga na proxima semana!."

        # Criar embed final formatado
        embed_final = discord.Embed(
            title="✅ DG BENEFICIENTE FINALIZADA",
            description=f"*{penalidade_texto}*",
            color=0x00ff00
        )

        # Separar por função para melhor organização
        tank_info = ""
        healer_info = ""
        dps_info = ""

        for membro, info in pontuacao.items():
            funcao = info["funcao"]
            pontos = info["pontos"]
            
            if funcao == "TANK":
                tank_info = f"🛡️ **{membro}** → +{pontos} pts"
                adicionar_pontos(membro, pontos)
            elif funcao == "HEALER":
                healer_info = f"💚 **{membro}** → +{pontos} pts"
                adicionar_pontos(membro, pontos)
            else:
                if dps_info:
                    dps_info += f"\n⚔️ **{membro}** → +{pontos} pts"
                    adicionar_pontos(membro, pontos)
                else:
                    dps_info = f"⚔️ **{membro}** → +{pontos} pts"
                    adicionar_pontos(membro, pontos)

        # Adicionar campos organizados
        participantes_texto = ""
        if tank_info:
            participantes_texto += tank_info + "\n"
        if healer_info:
            participantes_texto += healer_info + "\n"
        if dps_info:
            participantes_texto += dps_info

        embed_final.add_field(
            name="👥 Participantes e Pontuação",
            value=participantes_texto,
            inline=False
        )

        resumo_texto += f"**Tank/Healer:** 2 pts cada | **DPS:** 1 pt cada\n"
        
        if tipo_conteudo == "PONTUAÇÃO":
            resumo_texto += f"**Caller:** ⛔10 pontos (tipo PONTUAÇÃO) 📉"
            resumo_texto += "\n*pontuação atual:* " + str(obter_pontuacao(caller_nome)) + " pts"
        else:
            resumo_texto += f"**Caller:** Não recebe pontos"
        
        embed_final.add_field(
            name="📊 Resumo",
            value=resumo_texto,
            inline=False
        )

        embed_final.set_footer(text="Pontuação registrada no sistema! 🎉")

        # Enviar embed final e remover view da mensagem original
        await interaction.response.send_message(embed=embed_final)
        
        # Desabilitar todos os botões da mensagem original
        for item in self.view.children:
            item.disabled = True
        
        embed_concluido = discord.Embed(
            title="✅ PONTUAÇÃO CONCLUÍDA",
            description="As funções foram definidas e a pontuação foi registrada.",
            color=0x888888
        )
        
        await self.view.interaction.edit_original_response(embed=embed_concluido, view=self.view)
        
        # Limpar dados específicos deste usuário
        if user_id in conteudos_em_aberto:
            del conteudos_em_aberto[user_id]
            
        # Fechar todas as mensagens ephemeral individuais
        for membro, msg in getattr(self.view, "mensagens_ephemeral", {}).items():
            try:
                await msg.delete()
            except Exception:
                pass

class FuncoesEquipeButton(discord.ui.Button):
    def __init__(self, membro, view):
        super().__init__(label=membro, style=discord.ButtonStyle.secondary)
        self.membro = membro

    async def callback(self, interaction: discord.Interaction):
        if interaction.user != self.view.interaction_user:
            await interaction.response.send_message(
                "Apenas quem usou o comando pode definir as funções.", ephemeral=True
            )
            return

        # Menu para escolher função
        options = []
        if not self.view.tank_set or self.view.roles[self.membro] == "TANK":
            options.append(discord.SelectOption(label="TANK", emoji="🛡️"))
        if not self.view.healer_set or self.view.roles[self.membro] == "HEALER":
            options.append(discord.SelectOption(label="HEALER", emoji="💚"))
        options.append(discord.SelectOption(label="DPS", emoji="⚔️"))

        select = FuncoesEquipeSelect(self.membro, self.view, options)
        msg = await interaction.response.send_message(
            f"Selecione a função para **{self.membro}**:",
            view=select,
            ephemeral=True
        )
        self.view.mensagens_ephemeral[self.membro] = msg


class FuncoesEquipeSelect(discord.ui.View):
    def __init__(self, membro, parent_view, options):
        super().__init__(timeout=30)
        self.add_item(FuncoesEquipeSelectMenu(membro, parent_view, options))


class FuncoesEquipeSelectMenu(discord.ui.Select):
    def __init__(self, membro, parent_view, options):
        super().__init__(placeholder="Escolha a função...", options=options)
        self.membro = membro
        self.parent_view = parent_view

    async def callback(self, interaction: discord.Interaction):
        escolha = self.values[0]
        
        # Resetar flags de tank/healer se necessário
        if escolha == "TANK":
            # Se já existe um tank, transformá-lo em DPS
            for m in self.parent_view.membros:
                if self.parent_view.roles[m] == "TANK":
                    self.parent_view.roles[m] = "DPS"
            # Se este membro era healer, resetar flag de healer
            if self.parent_view.roles[self.membro] == "HEALER":
                self.parent_view.healer_set = False
            # Definir como tank
            self.parent_view.roles[self.membro] = "TANK"
            self.parent_view.tank_set = True
            
        elif escolha == "HEALER":
            # Se já existe um healer, transformá-lo em DPS
            for m in self.parent_view.membros:
                if self.parent_view.roles[m] == "HEALER":
                    self.parent_view.roles[m] = "DPS"
            # Se este membro era tank, resetar flag de tank
            if self.parent_view.roles[self.membro] == "TANK":
                self.parent_view.tank_set = False
            # Definir como healer
            self.parent_view.roles[self.membro] = "HEALER"
            self.parent_view.healer_set = True
            
        elif escolha == "DPS":
            # Se era tank ou healer, resetar as flags correspondentes
            if self.parent_view.roles[self.membro] == "TANK":
                self.parent_view.tank_set = False
            elif self.parent_view.roles[self.membro] == "HEALER":
                self.parent_view.healer_set = False
            # Definir como DPS
            self.parent_view.roles[self.membro] = "DPS"

        # Atualizar o embed principal
        await self.parent_view.update_embed(interaction)

# Função para converter valores abreviados (17M, 5K, etc) em números
def converter_valor_abreviado(valor_str):
    valor_str = valor_str.upper().strip()
    
    # Se já é um número normal, retorna diretamente
    try:
        return float(valor_str)
    except ValueError:
        pass
    
    # Mapear sufixos para multiplicadores
    sufixos = {
        'K': 1_000,
        'M': 1_000_000,
        'B': 1_000_000_000,
        'T': 1_000_000_000_000
    }
    
    # Verificar se termina com sufixo conhecido
    for sufixo, multiplicador in sufixos.items():
        if valor_str.endswith(sufixo):
            numero_str = valor_str[:-1]  # Remove o sufixo
            try:
                numero = float(numero_str)
                return numero * multiplicador
            except ValueError:
                raise ValueError(f"Formato inválido: {valor_str}")
    
    raise ValueError(f"Formato não reconhecido: {valor_str}")

# Função para formatar números grandes em formato abreviado
def formatar_valor_abreviado(valor):
    if valor >= 1_000_000_000_000:
        return f"{valor / 1_000_000_000_000:.1f}T"
    elif valor >= 1_000_000_000:
        return f"{valor / 1_000_000_000:.1f}B"
    elif valor >= 1_000_000:
        return f"{valor / 1_000_000:.1f}M"
    elif valor >= 1_000:
        return f"{valor / 1_000:.1f}K"
    else:
        return f"{valor:.f}"


# Funções para API do Albion Online
async def buscar_guilda_por_nome(nome_guilda):
    """Busca informações da guilda LOUCOS POR PVE usando o ID fixo"""
    # Se for a guilda LOUCOS POR PVE, usar o ID fixo
    if nome_guilda.lower() == "loucos por pve":
        return await buscar_guilda_por_id(GUILD_ID)
    
    # Para outras guildas, tentar buscar na API oficial
    return await buscar_guilda_api_oficial(nome_guilda)

async def buscar_guilda_por_id(guild_id):
    """Busca informações da guilda pelo ID na API oficial"""
    url = f"https://gameinfo.albiononline.com/api/gameinfo/guilds/{guild_id}"
    
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url) as response:
                if response.status == 200:
                    data = await response.json()
                    return {
                        'id': data.get('Id'),
                        'name': data.get('Name'),
                        'founder': data.get('FounderName'),
                        'founded': data.get('Founded'),
                        'alliance_tag': data.get('AllianceTag'),
                        'alliance_name': data.get('AllianceName'),
                        'kill_fame': data.get('killFame'),
                        'death_fame': data.get('DeathFame'),
                        'member_count': data.get('MemberCount'),
                        'source': 'official'
                    }
                return None
        except Exception as e:
            print(f"Erro ao buscar guilda por ID: {e}")
            return None

async def buscar_guilda_api_oficial(nome_guilda):
    """Busca guilda na API oficial (método menos confiável)"""
    # A API oficial não tem endpoint de busca por nome
    # Este é um fallback que pode não funcionar para todas as guildas
    print(f"Tentando buscar guilda '{nome_guilda}' na API oficial (limitado)")
    return None

async def buscar_membros_guilda(guild_info):
    """Busca os membros de uma guilda usando a API oficial"""
    if not guild_info:
        return []
    
    guild_id = guild_info.get('id')
    if not guild_id:
        return []
    
    # Usar o endpoint correto para membros da guilda
    url = f"https://gameinfo.albiononline.com/api/gameinfo/guilds/{guild_id}/members"
    
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url) as response:
                if response.status == 200:
                    data = await response.json()
                    return data if isinstance(data, list) else []
                return []
        except Exception as e:
            print(f"Erro ao buscar membros da guilda: {e}")
            return []

async def buscar_membro_por_nome(nome_membro):
    """Busca informações detalhadas de um membro específico da guilda"""
    url = f"https://gameinfo.albiononline.com/api/gameinfo/guilds/{GUILD_ID}/members"
    
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url) as response:
                if response.status == 200:
                    membros = await response.json()
                    # Buscar o membro pelo nome (case insensitive)
                    for membro in membros:
                        if membro.get('Name', '').lower() == nome_membro.lower():
                            return membro
                return None
        except Exception as e:
            print(f"Erro ao buscar membro: {e}")
            return None

async def safe_defer(interaction: discord.Interaction, ephemeral: bool = False) -> bool:
    """Tenta deferir a interação de forma segura sem exigir imports extras.

    Retorna True se defer foi bem-sucedido, False caso contrário.
    """
    try:
        await interaction.response.defer(ephemeral=ephemeral)
        return True
    except Exception as e:
        # Não importar tipos específicos — inspecionar nome da exceção em runtime
        ex_name = e.__class__.__name__
        if ex_name == 'NotFound':
            print(f"[safe_defer] NotFound ao tentar deferir: {e}")
        elif ex_name == 'HTTPException':
            print(f"[safe_defer] HTTPException ao tentar deferir: {e}")
        else:
            print(f"[safe_defer] Erro inesperado ao tentar deferir: {e}")
        return False

async def bloquear_comando_no_canal(interaction: discord.Interaction, nome_canal_bloqueado: str) -> bool:
    # Verificar se está sendo usado no canal proibido
    if interaction.channel.name == nome_canal_bloqueado:
        embed_erro = discord.Embed(
            title="❌ Canal Bloqueado",
            description=f"Este comando não pode ser usado no canal **{nome_canal_bloqueado}**.\n\nPor favor, use este comando em outro canal do servidor.",
            color=0xff0000
        )
        await interaction.response.send_message(embed=embed_erro, ephemeral=True)
        return True  # Canal bloqueado
    
    return False  # Canal liberado

async def permitir_comando_apenas_no_canal(interaction: discord.Interaction, nome_canal_permitido: str) -> bool:
    """
    Permite o comando apenas no canal especificado, bloqueia em todos os outros
    
    Args:
        interaction: A interação do Discord
        nome_canal_permitido: Nome do canal onde o comando PODE ser usado
    
    Returns:
        True se o canal está bloqueado (comando não deve continuar)
        False se o canal está liberado (comando pode continuar)
    """
    # Verificar se NÃO está sendo usado no canal permitido
    if interaction.channel.name != nome_canal_permitido:
        embed_erro = discord.Embed(
            title="❌ Canal Restrito",
            description=f"Este comando só pode ser usado no canal **{nome_canal_permitido}**.\n\nPor favor, vá até o canal correto para usar este comando.",
            color=0xff0000
        )
        await interaction.response.send_message(embed=embed_erro, ephemeral=True)
        return True  # Canal bloqueado (não é o permitido)
    
    return False  # Canal liberado (é o permitido)


async def atualizar_ranking():
     for guild in bot.guilds:
        # print(f"[GUILD] Verificando guilda: {guild.name} (ID: {guild.id})")
        if guild.id == 1183472048228548668:  # Substitua pelo ID da sua guilda
            print(f"[GUILD] Conectado à guilda: {guild.name} (ID: {guild.id})")
            await guild.chunk()
            # --- Ranking tags ---
            ranking = obter_ranking()
            tags = ["Ranking 1🥇", "Ranking 2🥈", "Ranking 3🥉"]
            top_nomes = [nome for nome, _ in ranking[:3]]
            # Garantir que as roles existem
            for tag in tags:
                if not discord.utils.get(guild.roles, name=tag):
                    try:
                        await guild.create_role(name=tag)
                    except discord.Forbidden:
                        print(f"[ERRO] Sem permissão para criar o cargo '{tag}' na guild '{guild.name}'. Verifique a hierarquia e permissões do bot.")
                    except Exception as e:
                        print(f"[ERRO] Falha ao criar o cargo '{tag}' na guild '{guild.name}': {e}")
            # Atualizar roles dos membros
            for i, tag in enumerate(tags):
                role = discord.utils.get(guild.roles, name=tag)
                nome = top_nomes[i] if i < len(top_nomes) else None
                for member in guild.members:
                    # Se o membro é o top i, garantir que tem a role
                    if nome and member.display_name == nome:
                        if role and role not in member.roles:
                            try:
                                await member.add_roles(role, reason="Ranking de pontuação")
                            except Exception as e:
                                print(f"[TAG] Erro ao adicionar role {tag} para {nome}: {e}")
                    # Se não é top i, remover a role se tiver
                    else:
                        if role and role in member.roles:
                            try:
                                await member.remove_roles(role, reason="Ranking de pontuação")
                            except Exception as e:
                                print(f"[TAG] Erro ao remover role {tag} de {member.display_name}: {e}")


if not token:
    raise RuntimeError("Token do Discord não encontrado. Coloque seu token em TOKEN.TXT.")
bot.run(token)

